import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
import matplotlib.pyplot as plt
from datetime import datetime
import json
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Configuración de página ───────────────────────────────────────────────────
st.set_page_config(
    page_title="Sistema Académico · Fuente de Gracia",
    page_icon="✝️",
    layout="wide",
    initial_sidebar_state="expanded"
)

SHEET_ID  = "1biSSyFbRPv3JNCxaGoHzewxVEd7S80tm4yGAbA9jpEo"
SCOPES    = ["https://www.googleapis.com/auth/spreadsheets",
             "https://www.googleapis.com/auth/drive"]
LOGO_B64  = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAC0ALQDASIAAhEBAxEB/8QAHQAAAQMFAQAAAAAAAAAAAAAAAAYHCAECBAUJA//EAFcQAAECBAQDAwQLCgoHCQAAAAECAwAEBREGBxIhCDFBE1FhFCJxgRUWGDJWkZShsbTRFyM4QlJVcoTS0yRXYmR0gpOVpcE0Q0RGsuHwKDM2N1NjdpKi/8QAHAEAAQUBAQEAAAAAAAAAAAAAAAECAwQFBwYI/8QAMREAAQMCBQMDAgYCAwAAAAAAAQACAwQRBRIhMUEGE1EUMmEHcRciU5Gx0RUWUoGh/9oADAMBAAIRAxEAPwCIkEHoEXEG2oJITewJ+iNpV1bcWAsBvzggBIvY8xYwEEcxzECEQdL7WvbnFyFFKwQEkpN7EBQ9YMW3JN+/ugQrkkC9wTcW2NuYgGmyrg3ttbvuOfqvFsECEQXud+V7mwggBIIUCUqBuCDaxhEK5QIWUq0hSdja1tvRsYt2tt64BaD1XhUIJNim50k3tfaDl3ei0XLSpJKVJIUkkEEWItzBHfFu4sd94EIg5agRvy9EAJG4JHMbH4xAALEkj0d8IhVFzZIud9gIrcgmw6Hb1QGyleanSL2Avf6YpckKKhqUbbnmP+toVCoTe1wOVthaDck87/8AW8EXbpUQU7puCDfYwIVvS8XAXBNx5u9j1i2CBCDcnkPiEEG3fBAhEBJ5E7DpB1tBAhBJJuSSe8xNnhkyLytxlkjh/EmJMLmeqs55T27/ALITTevRNOoT5qHEpFkpSNgL2ubkmITkpJ80ECw5m/TfoOt/+fOOkXBh+DVhQf0z64/FWqcQ0WT2AXR7mPJD4En+9Jz97B7mPJH4E/4rOfvYeKwgtFDuO8qSwTOe5kyP+BP+Kzn72Ke5kyP64I/xWc/fQ7FQnJSnybs5PTLMtLNJK3HnVhKEADclR2HpMR1zO4k0NuO0vAMqJl3dAqEwklBPK7aOajfkTt4GLlLS1FU60f7qCWZkW6VFQ4dcgadLqmKhhWXlGUi6nHqzNoSm3UlT1hDZ4nonCRRnFMM4fdqj6DYpkJ6ecTfl7/tQg+pUJXEGFceVijrxdmRW36ZIEgo9kVFT7qjchLTAtYkDkQmw32AJhK4KwJiHG9Xdk8LU6YmJdCgFzL4CENg8itQ825G+kXPcDa8eipMFpspfPNoN7bfus+WtkJsxqUU7N8OqSoSeUdRdHQu1qZRf1B5Vow2Z7IYrAeycmAnqUV+bJ+dwQ6UhwwiVpL07X8TOdo0ypZZkmQBsCbFa7k+pIhuMiMuJDMesVSnTtRm5Iysul1tbASbkqsQrUDceAsfGLUcGEOje9pJDd91E6SrDgDytlSHuF951IqOWlXkd/fiozTqR4nS/f4gTDk4Syw4WMVACi0uSeeP+oXV5xt0X6aFOhXxAwh8dcOOLKJLuTdCmWa6ygEqbQnsn7C52TchR8AbnoIbTCuHpar1U0ebqyaLVu07NhM60UNFd7BCl821X2F+ZsLg84nYXQVMRkp5SE8VVRE6z2qWKeGXI5VrYKB9FVnP3sVPDHkh8Cf8AFZz99DKyGM83snak1J15mZm6YTpQ3OKLrDgH/pPC5SbdN7XF0xIbKrN3C+P2vJ5R8yNUSm65F8gL/SQeSx6Nx1AjAqcOnhbnaczfIV6KqY/QixWk9zJkf8Cf8VnP3sHuYsj/AIE/4pOfvYeK8HrjL7jvKt5Qmc9zFkh8CT/es5+9iEXEphqi4QztxDh3Dsl5DTJNcuGGO1W5oCmG1q85ZUo3UpR3Jtew2sI6gRzX4xvwkcW/pyv1VmLNK4l2pTXAAJooICQeQAgttGgokQQQQIRAuxUq1wL7Am/zwbXNtvTFyElSkpSCpRNgLc77C0CELUVaTYCwAuOtup+j1R0h4L/wa8Kfrn1x+ObnM7C1zsI6R8F34NOFP1z64/FOr9oT2bp37de7lGoxbiKk4VoMxWa1NolZOXF1KJ3UTySkfjKJ2AG5MbGcmGZSWcmJhxLTLSCtalGwSkAkknuHP1RDnHeIMQ565ls4fw+FIpDDpEsDfQhANlTDlupB2HQEAbkkrh9F6l93GzRuVHUT9sWG5VMT4mxznxi80OhsuMUhshSJckpbaRyDj6he6ieQ3A5AEgkvbg3LjBOUeF38RVbs5yelWu1mKg+i5TYe9aTuE35ADcki57lxlxgqjYFw61SKOwBYBT76kjW+u261EdfDpyEMpxA1Wo5g5kUzKigOlDDbodqLwuUpVa+46hCbmx5qUBtaNE1XqX9iEZYxv9vlVu12xnfq4pPYfouIeILHjterKnpHCsi6UNtJJFk7Hs0dCsixUscrgDoBKPD1FpdBpbNLpEkzJyjKbJabTZI/5+JjywjQabhjD8pRKUwGZSVaCGxsSe8kjmoncnqSY3EZlZWGYhjdGjYKzBCGC53WFWm0O0qaZcX2aVtKSVfk3B+i8RQyer2Fss67UZ1qaqtVVMshggyrbaU6VX1DzzcGJHZu1UUbL+rTwVpWlhSUEflKGlPzkRDEX6xi1eKS0rO3GdHbroHSfTNNizXyTjQbKSJ4hMPk70mo27tKP2oMyct6Lm1g+XxLSpUUutvMB5h5aQC6CLhDtuYI67kX6i4iOMu0qYmWpdsXW6sNpAFySohIHxkROfDckin0KTkm02S0ylAHgAAPogwnEJ8+cHZR9Y4DQ4c1jIhqUwWQuM3KouayozFk23p6WKmmBNpCu1SkbtL1e+UkC4V+Mn0XOlzjyGnqDMLxTl6uZKGD2qpFpSi8yRvqZUN1D+TzHQm9gquKbA7pk2MxcPBUrWKQQ5MOs7KU0k3Ss+KCL3/JKgdrQ52UeLmMcYFkK4nSHlo7OabHJDqdlD0X39Fo9Y6rfEBUw+07ji650IQ4mN2/Ca7h6zxTiBcthfFzyWqrsiVmz5qZkjYJV+S5/wAXgdjIIEEXiNfE3lEFIdx1hRlTUyzd2oSzIsV287tkW5KSRcgcxvsQdSq4Zc0jjOimh1mYSquSDYuokBUy0LAOem5AVbqQesQ1lLFNF6mnGnI8J8ErmO7chT1iOa3GLf3SGLD/AC5Uc/5qxHSkRzW4xfwkMW/pyvX+asxn0vvVx+yaO/O8V2087qJ5W5CKKsTcDSLcr+EEaKiRpP5MEEECEG42/wA4Lm4IsCORG0Fzp02Fr928ANvXAhXECydrncne+3+UdIeC8k8NeFD/AEz64/HNwgjYi3pjo9wdPIl+GPDDzigEITOFRPIATj5irVC4ACew21Sd4w8cuU+jy2Cqc+tExUU9rOFCrEMXICDbopQ9YSR1hq8rqlmdgBqZVQsCTLrs4Qpb79OeWspA81KVAiyb726lVzGfl/JuZt8RUxWKjd2QZmDOOIO47FshLLfoJCLjqNXfEmc38QnDGA6hPMLDc0W+ylyNiFq81JHoJv6o1KmtjwymbTlgNxdyrUlFJiFQAw6k2CYioZz50yEm7Nz2DUS0s0m7jrtOeShIva5UVWAhD5eV7MqkVyo4xouFJiqTVWUS5NOSDriffqUoIKSAATa/P3oHSKjF+LPfe2es37/LHPtg9t+LSbnE9YJ8Zxz7Yw2dWUkbS1sI1XQfw0rH2cZUvfuuZ7fAdX91v/tQfdcz2+A6v7rf/ahB+2/FnwmrPy1f2we2/FnwmrPy1f2xD/s1H+gFJ+Gtb+qlpmDjXGNdy1lZfF9ORTJ2aqR0sBhbSiy2hJ1KSok2K1C3TzYbCMupVOp1NxK6lUJqcKPel90r033NtRNgbDlGJHlcQqm1UxkaLBdL6cwc4VRiFxuUp8qqaKnj+lNKbUttl3ylYSm5s2NY27yoJ9doWLubWeXaLDOBXENg+aDSnyQm+wPnc7Q19Pnp6nTPlMhOPyb2nT2jThQu3xg9Bt4RsvbdiscsS1j5Yv7Y0sIxWCiYWyR5rrzvVXSlVjNQJI32AS2ms0s8JqXdl5nAXatOoKFoVSHyFJIIIsVbixsRCGy2xnmJl5UJ3D1FoLpmJ5wPmnzEm4pSDpPnISFAi6bA3vcJHdF/tuxZ8J6x8sX9sV9t+LfhPWfli7fTG8zqylaws7IsV5N30yrCQe6l25mxnmtBSrAiilQ3HsU+R/xQ0TU3inA2NZTFrlAmKG6ZpTrbBl1sNKBN1toCt9JBtbewI7hCokcTY0np5iTlsR1hyYeWlttInF3JUQB17zD+5rYCerGRi6RNTD8/U6cx5WzMOqKnFOoBUdzc+cNSbdxjSwrqCGV5jbEA12hXmeoelpcJDTJJc7pyMJVuSxHh2QrlOcDkrOspdbPWxG4PcQbgjodo53cY34SOLf05b6qzEnOCzE5mqHVMKvuFSpJYmZdKj/q3CQoAdwUL/wBaIxcY/wCEjiz9KV+qMxDJT+mqnRrIik7kYKaMi1iesFgbm4v0G9zBBc3vfeJkqNu754IqbX5wQIVIICCDYgjvgNjawttzvzgQixvYC+3TeJz5XVs0HgOlJ9Cil1UrOsNm9iFOTzyAR4jVf1RBoKVpKQpWkm5F9ifREsDNLl+BfBUqnlNVR5CvQmamnPpQII4+5PG0+QmSvyxkpxOCSjdlh6t15SAFPzSZVCrblLaQT6rrt/Vg4rK+HKpIYebcH3pPlLwv33Sn1++hFZO56UvAGB5bDrmHZqddadcdW8h9KEqK1qVtffYED1QpZriUwvMul2YwI864ealvtlRtsLm14TGsJrKuZ+Vv8K50/i8GGysmeL5Uz+pPLUn44NSfyh8cLpzPxJx0ipIoLacNpY0KpnZMFwub+f2ltQHvdr9IUXujcI/xfuf2zf2RgSdG1jLfP2XRh9UoeI00epP5Q+ODUk9R8cO4OIzCNyfaA7/bN/ZGXN5oUTGWX2JZmn4X9iTJttNJcWtCta3ipIA0i9xYm8U6npmpp4zJJsFcoPqKysqGwMi1JTNc4L+EUEbXB8gapiul04AkPzSEK8U6gVH0BIUfVHno4zI4NG5XRqioEELpTwLqwUKuKAIo1RUm2xEsux8eUHsDXfzJU/kq/sh25riYw7ITT0mzhGafQwtTaXBMNgLCSQFAHobA+F48/dR0P4FzXylv7I9W3pSpc0ED+P7XJX/U1zHEdtNR7A138yVP5Kv7IPYGunlRKn8lX9kOx7qShfAua+Ut/ZGkxtxGU+u4UqNIkMOz1OmppgttTLc0lKmlnkoFIB2PcQYkj6QqXOAI3+39qP8AFBwGsa3PDjgecexG7X6tJPy7MkNDCHmikqcUN1AEdB86j3RJFaApopIuCCCO+IpZecRrGG8ISNGqlFqNTm5ZJSuaVNgqcuokElVySAQN78oUXuq6Sf8AdGeH62j7I0oun6qlORrb2+QvGYt1CMVnM0htfhIbKsKwFxRzFCF0Szs4/IBPe04C41t4aUer0wz/ABik+6QxYf5cr9VZheVLGUtizPqmYsk5JyRS/Pymptawo3SpKTyABuEiEFxgk+6MxWb7apX6ozGzicTmvjc/ct/hYlG4FrgNrppLEcxBzUbA/FAQNrG+2+3KC4HID1i8UQraLDwggghUI3Atfb0RcRc3FwOkX6FlCgixCBqJAAIuQncnfmQLRbpGwAUkW9N/RBZOVAkm4O59ESqmmlngjwEsAlLdYf1W3sO3nACfXYeuIspJCgoE3B23NxbrEzMNUlVU4CKb2SVKXJqmJtNvONm6g8Vf/nVvD4XZJ4z8hQzC8ZTpcNlGoVQyZoMxMUqRmH9DqXFuS6FKJDqxubHwhyPathw3vQ6b8mR9kNFwY1UTeWs3S1Hz6fUFgC/4jlnAf/spY9UPrFLEXSMqni53TqZrXRjRRxRSKS3xdv05dNlDKO0lKksllOgeYNwm3PzTvD6e1TDf5hplv6Kj7IZjHB9iuLvCs8pQCKjTuwudhcF4esklIiQI5Xh9dM8tjIcfaE2CNt3C3K0vtUw1z9gqZ8lR9kMlxQuU+myVKoVMkpWUS+6uZeSy2G9RQnSNWkC+6+vdEhibc4iXxGVMVDMuYZSrUiTZS1YG4BsVH5lCMDEqh4gIJ3XtOjKJs+JsNtG6puIczhtpAqWYwm3EakSEupwX5BSrJHzFUNnEheFCnobpVTqSrBb73Zpv1CALn0XXaMPD4884XUOr6r0+GPtzoneOFcNk3NBpl/6Kj9mKe1PDX5hpnyVH2Ru7QWj1YmeOVwUsadwtJ7VMNfmCmfJUfZCRzkw/QJTK3E0yzRqey4imPlC0S6EqCtCgLEJ2NztaHI+iG34lJ5Mjk1XiVaS80lhJ8VrCfovFilkkdMwXO4UUrGhh0Wl4bcP0abyhpExOUmSfdX2pUtxhKlH74obkg32EOT7VcN/mKm/JkfZCdyCklSOUGGmlI0qVIocI/Suq/rvC6J82Fq53undZx3SRRtyC4UO875SUb4kKVIU+WZl0JckklDKAhNy5e9htfcQzXGCf+0bisX/GlfqjMPK457cOL4dl98ZZqwSLG40yyPOsR01Iv64ZzjEJ90biwEkgLlbC/L+CMxs1rjaFh3DVVpgLvI2umjIO1xbaA7AW5233g3JtBFRWlW0EUggQvYlKgkhOkgWNr7787kn/ACEVCQTYG55A3+nuiiLGxvzNjcbfN9kegSU7AAEc78x4QqUqjd7DYi3Mg846G8LdNarHCjRKXMWU1Ny8+wvuKVTT6T8xjnqLGwtbbnHRng2B9zdha4H+2X+WPxVqXZQCOCnBoNwU0fCpV14VzWqmDqistGb1sWUbff2VKNt+pTr+KJca0nqL+mIjcXOEHqBjWWxnSw4w1UrB5bRKS3MJTbVcWIKkgHv81R6x5YByrxzjTDMrX6RmKksTAN0LmJjW2sbKQoBXNJ+PYjYiNaspIatjasvyg6HS+qzopnwuMWW6cDi2KqPOYLxrLo1rpFSuvTzKCUrsf7Mj+tD7yM0xNyjUyw4lbTqAtChyUCAQYilirILMRqgTkzNYtaqyJZlTwlCt5ZcKUk6U6ifOIBAv1MaHKbLrGOP8NKqlFxyZRph3sFy7j72popAIHmq2BSUkW9HSI30VNJTNImH5dDp5TmzSNkP5d1MudfSzKPOlQ0oSSfCIOYjqSqziGo1VRJ8rmVup3vZKiSm3gE6R6oc5GXGLsu6bVcTV/GHskyinusMsIeeVd1wJQkqC1WsASe+9oaIAAWA2jxGPBkbmxxuuuxfTmmLmyVLhrsqE2F4klkar2MrEhhzXYsYeROvp7nJh5SrW7wlI9UR5ockqpVuRp6b3mZhtuwF/fKA/zhTs4fr+aGcWLRh2vJpHkTgYSvUtIW00Q0m2g8vvd9++8WemqFk4kkebADdN+pOIGNsUDedVMjWn8oQa0flJiL3ufcy/4wk/20x+1B7n3Mv+MNP9tMftRv8Aoab9YfsVyjvy/wDBShK0flJ+OGH4xJ5x/C9DwrJqKpus1JKEJAJuE7ch/KWiEv7n3Mr+MNG3/vTH7UN1R8vcUYnzPnMJS2JxNzlIBWqoOOOKQyUqTcJuSoHUoC4tuCekXaGgpw8yd4EN12VeaeQjLk3U2KJKMUykSVOYIS1KsIZQO5KUhI+YRrMw8QsYYwXVK464j+CSylpBPvlWskeskCGA+4BmZz+6KP7eY/ahpMz6LiPDuJPajPYnfrswoI1stPOKTrURpQUqUbq96bW2uIKPCoKmX8soNtTpwiWpexli1OjwY0FyexRWsXzYK1MtmWbWQTqccUFuKv37C/6RhguMT8JHFv6ct9VZid+SeDk4Iy9p9HWEmcKe3m1J5F5dir1DZI8EiIIcYtvdI4sPcuWvt/NWYrVNQKirc9uw0CngjLIgE0fjBuRbaCAAnlDlIqkJvzPxQRSCBCyElQChe2rc+MegAIuRqJ6xRKd9J27t9vj6x6pTa5F9PXx/5wqcq6NVyLJUTcC4AEdFuDj8HHC29/8AS97W/wBsfjnaOSTtpvv4/HHRPg7sOHPC9r2/hfP+mPRUrPaErN0ucw8K07GWE53D9SRdqYR5iwPObWN0rSehBsfHkdiYitlNiyqZL5kTuF8TpcRTHntE0ACUoNrImEDqki17blPeU2iZh6Q2Ge+VknmHREuSxbla5JoV5JMKTsrr2ayN9JPXfSTex3Blw6sYwGCb2O/8+VBUxEnuM3CcaUmJedlG5mXeQ6w8nUhaFBSVJI2II5gg3iNGHpj7jnENPUeaUWMPYiUFMqOyEFSlFBv00qKknuCgTtaE7lBmlW8ra0vBuNpeZFMZd0FKgS5JXPvkj8Zs87Jvzum/IvRnNhGmZs5dIm6FMS85OS6VTFOfbUkhZIGpvV0CgLG/IgX5RP6c0chZJrG7n+FH3O8AW6ELVcVtX7DDlOpCF2VOP61AdUNi59WopiOULDM6bqrgw5Sqy++7PU+jsomO2N1pdUnUtKv5QGkE8zp33hHxz7FSDUuAOgX0N0bSenwxhO7tUtMlKe5PY6bmEAWp8s7OG4uLpQQn51A+qNzwVLU5jXETqzdSpRtSiepLhJPxmN/w20wewOKawtA1GXLCCeYsgqVb0kj4oT3BL/4wr/8AQ2v+OPYYBGG4XKfsuU9dVhnxgN4bopZbRSwivdGsr1Xp1EpUxU6nNNSsrLJK3HXFAJAHj39LdTtzimGlxsAvPEgC5SYzqxszgTAk7Vwpszq0lmRbUffvKB03HcPfHwBhH8KWEXqLg1/ElUBVVK855Qtax53ZC5QCTvc6lK/reENzIeyHEDmwmZmmnWcH0ZWzStgpJINlfy3LC/VKRbxLw5uZrYdy5pPkrXZTdV7MIlqeyoDT0Cl294kejfkB3bToHRximjF3u3/pURIHOMjtgsjPLMiSy9wut4Lbdq80kpkZYkEqVy1qHPSkkXPXYDcwzHDBgCdxPiR7MbE5dfbQ8pcqXR/pL5USpw36JNwLbEnppsdFlzgjFGdeM3MV4qdfRSAv769ukOAHZlofki5BI5b7kmJfUqQk6XTpenSEs1LSku2ltlptOlKEpFgkAcgLRLPIzD4TBGbvO58fCSNrqh+d2gCzQLbRzW4xvwj8WWBF1y1/H+CsR0pEc1uMW54kMWDc+fK2H6qzGVS+9XX6BNLclBTZOxvewv3Wv/lFvjARa+x9YgIF+oEaKiRYwRUrWLAKOw74IELLbCbpudJvubx62USoE9dze/zxYGlJ0BxKkggKFwQSDyIvzEeyEqBSgtq1KsU7EXB5W74cUpGqq2lVhcmyuY5x0U4PPwdML/rf1x+OeKAnbUCLX8b90dDuD2/udcL/AK39cfilV+0JzN07sEEEUFKkDmvljh7MKnhuotGXn2gRLzzSR2jfXSb++TffSfTsd4j/AIWwrmrlPmNTqZIuOvUeoTzbLj7aVOSjqFKGpS0c21hNzfblsoiJdx4TLKX5dxlSiErSU3SbEX228YvQYlLFGYjq08FQenYZA88KE2Pan7M40rFRCipLs2oIJ/JTdKfmSI0hPjEqmMi8CtpIMnNO9xVNuX+YiEriXJ1+TWtVBwnQKm30RM1SbaWfDYkE+sR5f/FyTylxIAK6/T9c0NLTNiaw6Cy3uTUnL0vJ4tuutomZ2Wdmi2VALKVX0mx3tYAX8IargrcQ1i7EC3FpQnyJu5J2Hn98J/MnAWbdYryak/gl6UQ3LtyrDFPmEOobabFkpBCyojmd994SUvlfmaF6GcH1pGvY+alKVfpEqAt6Y6PhuG00dE6IzAF1v+lxnFcQkq6104ZuSpe4/wA38EYRlnDNVducnE+8k5Mhx1R6XsbJHiogRF/H2Y9RzTraGK7VGKBh2XV2nYjUs7HY6Ui7jljysEjn3x6UXh+zOqKx21JlKagkefNzaOXXzUFW/ptDkYW4W2UlDuJcRrd5FTMk3oB7xrVc+sJBiWGLCqBpJkzO8qm91VOdrBIr7r03IUJnBeVVCfprBGkTCkB2cfUffL0pB849/ndLWsLKbKrh9qVWm04gzHee++q7UySnCp50ncl1y99+oBJ7yOUP7gfAGE8GSvYYfo7Esojz3yNbrh71LVdR9F7DoBCqsO6MifFmtBbTNy33PJVqKjJ1kKxJCSlafJNSUlLtS0uykIaaaQEoQByAAsAPARmc4OkHSMQkk3KvgACwVY5q8Y1/dIYs521yv1RiOlUc1uMWx4j8WW56pW/yVmLNL701+yaQgpURfcHmCDy7iPpi23MC5PgLwQbW57xoqJEEEECFnJUSU6lKV0593SNtQqDV6xK1GbpskuZYprPbzZSQOzQVBAUUkgm6lJFkgnflGnaSVAqANk8z0h6uGcgyONyQLexsrcfrrMMkcWtunDUpLVbK7H9JmfJKjhmblXTIvVAB0oTeXaALq76rXSCklN9QuARuIXWX+IuIihZYSM1g5yabwm06WZRTUpJu3ccmSgpSFpU6oqecIAIO/La0OziuWqDNcxHMLrTU5TXpDF3k8qJPQuVdTo7QFwk9oFXSRsAOW8NfLzCJXJ7JmcpEnVJ7FsvU55+jS7CUFh4pnEqcS4CdZUSlASEi1iq9torOmLxqEuW2yzMQ5k8UeHlVE1mpTMoKZLNzU2VU+RKUNOLLaFghshQKwU+aSQQQbWMarFGdfEXhedk5Kv4hdkX5yWbmpdC6bIqU40sqCVjS2QASkixsRbcCF7XqPMVTDeNHKfOzjtFxFSabMUmVmFFS5DtqkA7L26aHisbbWI3JBMafjUpzxreEqq+2w2rXO09BbWlYDTEwC0LpJAOh0XB3BBuAbwwFpOycb+VZPY04s5HEtOw7NTE6zU6klwyjJp8hpeDaSpelfZ6CUpFyNVxcbbi+DLZlcUM1jWawTL1Z9zEEqjtHpNMjT9SE6UqJ1FvSdlJOxJ3h6poMfdCvPKcKBi6rAabaw0aKCoIv4791+cRsyTNJONscGgmd9jParV/JDOhIfDfYnSXNHmhXK9ja/KBtiCbINxylS3mXxRuPiXTPzvbqqfsT2RpUmFib0FfZlJbBHmDVqI023vaPeUx1xYTdeqdCln6g5UaW2l2cYTS5H72lQJSQot6VFViUhJOqxsNjZ4/975MqP39WJQXr89ftYVfV1vy8YQ0manOT9Jw42ZuRla3hKhNSVaZukSNUaZW5KlShvZfnpIG+4tbmEu3wk18pB4PzV4nMYS1SmcNVWaqrdN0maSzTJHW3qCikaS2FKJ0KsEgnblyjZSWNuLGcpFKq0pOzDklVlobkXRI04B5SkqKQAUApuEqO4HKMHh3rkhhLBmPMQYpVU1exVapUy4mSKS66+h19QQoqIGlStlG5Niq17wsadPUSpZqZIVNSKg1iCap6X1NDR5ImWUZpQAPv+0C1EctOnxhziAbAJANEn8M414s8TSDc7QpyZnGVth0KEjTkXTrcbvZSAR57TgsQDdJ2sRekxjTi4YpZqa5mbMr5Z5DrRIU9Z8oD5l+zCUoKie1BTsLXF723hZVxEgceYTm6XpEjUMGVqpMhJ2QmZD74SP0e10+FoTeWwofsfw/9uaj7N+VzfkwT2Zliz5U9rKyTq16tGmwtbVfe0NzDwlt8rWoxtxcu4km8OtvVBVUlJZM08yKZIWS0onSoK7PSq5CgACbkKABtYY1DzF4rqzQmK1TZyemKfMTaZNl80qRQFPFwNhNlNggdoQkqICQrYkHaFbk6vFi6e+rEjaUsJquGU0hQLd1SXsrMaSrQSff9oLKsrlcWtCVzATXqtlfiWVBnaRO4fmpt9oAlDVXo788oqHcotvgHewsbbnktx4QsudxfxfydfkqDMTMyiozzLr0qx5DTVFxDQTrUCEEC2pOxIJvteEfN59cQknQafXpnFLjdNqa3W5R806RIcU0UhwABokaSpN7gc9r7w8+BZWcb4nMT16TQ0pcjh+js/fXEoSA6mTC/OUQAezQ6Bvck2FzYQzXEJShQstcNUfRpTI4mxDLJAHRMw2kEeBAB9EOjLS4AhBvbdX0TPPiNrVGq9ZpeJXZqQorSHKi8mmyIDCFkhKiC0FEHSr3oNrXNhDa5wt45XjiYn8xW3EYgn2Wpl1S0NIUtsoSltWluyU+ahIsADtuLkw83BTLyNRouPKJVN5GsCQpr4sLgP+UtpKe4hSkkeMJXjJqSa1mhS6w3bTP4dkZpNrAWWhStvCyokDg2SwCTW26afENAq+HZ9qRrUkuUmHWGphDZUlRU24kKQoFJI3SoG3MdQI1ndzI5CHT4nf8AzBpv/wAdpf1VENWdjaJ2Ou0Eph3RBAefKCHIWShW5F7A93WN3QsSVqgyVSlqVPqlWqowliaCUJKnEJWlYAUQSkhSQbpIO1r22jQAk99u8xeCT1NrbXgdqLFOHwnLfzyzSfqUpUXMTBc1KNvttKVISpCUvBAcCklspUVBtJJUCbgnmok0OdWZjtLTS04jQ1KpmhNpS1ISzakOh4PBSVJbCk/fEpNkkCwta1xDbFRJAA5+MezRsbHYKFrj6YZ22+EXTiv5x5jTlUcqL+IwqYdTLoUTJsJQAw8H2gEBASmznnGwGo7KuCRCeq2LcRVaneQVCoGYlRUHamElltJEy7YOLCgkHztO6b6Ra4AvFcCy1Imp6bFXLAQ3K62Q86lCVOdo2m27rQUdJWba08r72sdzOUDCzDc3NtVmUfQFTiGWBNAqGlD4auLAquW2lBQJSrWE2vzQNaDoEv3WVPZ1ZmT+J6diScxO49UqYHEySzKMpQ2HUlDhDYQEKKkm1ykmwG+wtrZPMzG0jjuaxrJ1hDFem2+yfmUyTJCkFKQR2ZRoAKUpFwkHa/MkkxLRsP0wGpys/JzkuKmpHkTE6lxSpfW4Ry85ICUJGolV9YII6+yMP4ZVNppya7KlSFWVNGbQhtzUlxSVaiSEpCS0FAJUoKKh6FDWgbJpuvJGaOP25tM0nE00Zj2WNZ7UtNlflfZdlr1FJJHZkp0HzdO2m0ZUnnJmZJ1up1mXxS8idqSWkzKjLMqSQ2kpbKEFBS2UgkXQEkXO+8Yc3hyiuUaZqEjVZZt9uVlnGZZdQaW664pLPapCQkEAKcXtcFPZqBBG4y3cOYdmy2puqUunoTJKJQufbU52oSSCpQWtK7kEAJKCeelO1yzPCWxSWlsQ1mVoNUoTE6U06rOMuzzRQlSnFtKUpCiogqBClqOxF+t9oU0hnHmLT6RR6XKV9DcrRVJVTh5BLksaUqSnzi2VKASpWyiQb3NzaPFugYSVNTkumsqShC1sNPuzLSRdDzKe2Cfx0KQ44oJBBPZqIJ5DClaXSWanOtpVT51SZBLrDE3UEJbDpdbSUKdQtKVkIUtQ0qHQEXBEGRp4SC4XnJ5hYwlfYwMVhQ9jJB6nSZUw0otyzuoON7pJIIWoAm5AIAIsLZtMzWzApmFJHDMjXyxSqa8l2TaRKMhbLiXe2SQ5o7T/ALzziCo35EEbQO4dws5NNkVxuXYWtvtSmYaX2KitKXWkgqClhIUpQc96QkC5JvHmih4TfQ4WayphSRLKImHm7/fglSgkhQuW7KSokWJtcJtuBrPCNfKz0Z1Zns4jmsRM4qdRUJqWRKvKEqzoDSFKUhKWyjQkpUpSgUgEFSje5JOofzKxu7gheC3a++uhrJK2FNNqUoFzWU9oU9oUlZKtOrc72jPk8J0KoyU1NSFRmFsMyUw+rUtvUwppL5SFpBurWWU2SndIXckgAq1k3Q6LK4ppUuKi1M0x6abbm3TMtlKUF1SSoqQohIUhOuxPm6rE7Akysvsg3RU8xsZ1EVQTlZLnssJTy0iWaSXhKABgXSgFOgD8W1zYquRCipWe2YTFRlXanVmKjIpqiahMyxp8qkvL7RLiwHOyKmyooAJT8RHNNzshQTiahIbfkPIJksCeMs+tLKCp5SVae1V2qAEJTcrtY3I80pJ1+KJWSZlZR5pmQlZxZWH5aTnBMNJSNOhaVBayCo9oCnUbaUnYKAJlaeEa+Vk0TH2LKLPVGYotTEkupTrM7MBthshTzLqnGinUklIStSjYEAg2II2iyv47xTXMPS+HqrU0zFNlUS6GWvJmklIYbU0yNaUhRCUKUNyb3ubkAwmTbUbbi3oi4EaCDqve4328YcWNveyS6UuJMf4sxFSvYus1NEzJ62HNHkrKDdhsst+clAVZKCU2vY8yCd4TBtc2BA6XMEEAAGgQq3gikEKheiANAPM3t80UWrfkB6oIIVO4VzZ5+EezW+kE9D9EEEBSleze6N4qN0iCCDhNcqXJB9H2RRPvvXBBC8ocqk7eqKd/ogghUBVJuQD4R5K2KfTBBCcpOFaoxbfeCCFSFXEqSHEhStIG4v76xvv3x5q3TfxgggRwrXDsn0RYBdVjysfoMEEBQFb1gSbJ12B35dIIIYUBVggggSoggggQv//Z"

# ── CSS Institucional ─────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600;700&family=Source+Sans+3:wght@300;400;600&display=swap');

html, body, [class*="css"] { font-family: 'Source Sans 3', sans-serif; }

.inst-header {
    background: linear-gradient(135deg, #0a1628 0%, #0d1b3e 60%, #0a1628 100%);
    border-bottom: 3px solid #c9a84c;
    padding: 22px 32px;
    margin: -1rem -1rem 2rem -1rem;
    display: flex; align-items: center; gap: 20px;
}
.inst-header h1 {
    font-family: 'Playfair Display', serif;
    color: #e8c97a; font-size: 1.5rem; margin: 0; line-height: 1.2;
}
.inst-header p {
    color: #a0aec0; font-size: .8rem; margin: 2px 0 0;
    letter-spacing: 1.5px; text-transform: uppercase;
}

.materia-card {
    background: linear-gradient(145deg, #0d1b3e, #152444);
    border: 1px solid #c9a84c44; border-left: 4px solid #c9a84c;
    border-radius: 10px; padding: 20px; margin-bottom: 16px;
}
.materia-card h3 {
    font-family: 'Playfair Display', serif;
    color: #e8c97a; margin: 0 0 10px; font-size: 1.1rem;
}
.materia-card .stats { display: flex; gap: 20px; font-size: .85rem; color: #a0aec0; }
.materia-card .stat-val { color: #e8c97a; font-weight: 600; font-size: 1.1rem; }

.metric-box {
    background: linear-gradient(145deg, #0d1b3e, #152444);
    border: 1px solid #c9a84c33; border-top: 3px solid #c9a84c;
    border-radius: 10px; padding: 18px 20px; text-align: center;
}
.metric-box .label {
    font-size: .72rem; color: #6b7280;
    text-transform: uppercase; letter-spacing: 1px; margin-bottom: 6px;
}
.metric-box .value {
    font-family: 'Playfair Display', serif;
    font-size: 2rem; color: #c9a84c; font-weight: 700;
}
.metric-box .sub { font-size: .78rem; color: #a0aec0; margin-top: 4px; }

[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0a1528 0%, #0d1b3e 100%);
    border-right: 1px solid #c9a84c33;
}
[data-testid="stSidebar"] * { color: #e2e8f0 !important; }

.stButton > button {
    background: linear-gradient(135deg, #c9a84c, #e8c97a) !important;
    color: #0d1b3e !important; font-weight: 700 !important;
    border: none !important; border-radius: 6px !important;
}
.stButton > button:hover {
    opacity: .9 !important;
    box-shadow: 0 4px 12px rgba(201,168,76,.35) !important;
}

.gold-divider {
    height: 1px;
    background: linear-gradient(90deg, transparent, #c9a84c, transparent);
    margin: 20px 0;
}
.versiculo {
    background: #0a1528; border-left: 3px solid #c9a84c;
    padding: 12px 18px; border-radius: 0 8px 8px 0;
    font-style: italic; color: #a0aec0; font-size: .85rem; margin: 16px 0;
}
.section-title {
    font-family: 'Playfair Display', serif; color: #c9a84c;
    font-size: 1.3rem; border-bottom: 1px solid #c9a84c44;
    padding-bottom: 8px; margin-bottom: 20px;
}
</style>
""", unsafe_allow_html=True)

# ── Conexión Google Sheets ────────────────────────────────────────────────────

@st.cache_resource
def get_client():
    creds_dict = st.secrets["gcp_service_account"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return gspread.authorize(creds)

def get_spreadsheet():
    return get_client().open_by_key(SHEET_ID)

@st.cache_data(ttl=30)
def get_materias():
    return [ws.title for ws in get_spreadsheet().worksheets()]

@st.cache_data(ttl=30)
def get_estudiantes(materia):
    try:
        data = get_spreadsheet().worksheet(materia).get_all_records()
        if not data:
            return pd.DataFrame(columns=["Nombre", "Nota", "Fecha", "Letra"])
        df = pd.DataFrame(data)
        for col in ["Nombre", "Nota", "Fecha", "Letra"]:
            if col not in df.columns:
                df[col] = ""
        df["Nota"] = pd.to_numeric(df["Nota"], errors="coerce").fillna(0)
        return df[["Nombre", "Nota", "Fecha", "Letra"]].dropna(subset=["Nombre"])
    except:
        return pd.DataFrame(columns=["Nombre", "Nota", "Fecha", "Letra"])

def _limpiar_cache():
    get_materias.clear()
    get_estudiantes.clear()

def guardar_estudiantes(materia, df):
    spr = get_spreadsheet()
    try:
        ws = spr.worksheet(materia)
        ws.clear()
    except:
        ws = spr.add_worksheet(title=materia, rows=500, cols=10)
    rows = [["Nombre", "Nota", "Fecha", "Letra"]]
    for _, row in df.iterrows():
        rows.append([str(row["Nombre"]), int(row["Nota"]), str(row["Fecha"]), str(row["Letra"])])
    ws.update(rows, "A1")
    _limpiar_cache()

def crear_materia(nombre):
    if nombre in get_materias():
        return False
    spr = get_spreadsheet()
    ws  = spr.add_worksheet(title=nombre, rows=500, cols=10)
    ws.update([["Nombre", "Nota", "Fecha", "Letra"]], "A1")
    _limpiar_cache()
    return True

def eliminar_materia(nombre):
    if len(get_materias()) <= 1:
        return False
    spr = get_spreadsheet()
    spr.del_worksheet(spr.worksheet(nombre))
    _limpiar_cache()
    return True

def nota_a_letra(nota):
    if nota >= 90: return "A"
    if nota >= 80: return "B"
    if nota >= 70: return "C"
    if nota >= 60: return "D"
    return "F"

# ── Exportar Excel con estilo ─────────────────────────────────────────────────

def exportar_excel():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    navy_fill  = PatternFill("solid", fgColor="0A1628")
    gold_fill  = PatternFill("solid", fgColor="C9A84C")
    accent_fill= PatternFill("solid", fgColor="E8C97A")
    white_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    navy_font  = Font(color="0A1628", bold=True, name="Calibri", size=11)
    gold_font  = Font(color="E8C97A", bold=True, name="Calibri", size=14)
    border     = Border(
        left=Side(style="thin", color="C9A84C"),
        right=Side(style="thin", color="C9A84C"),
        top=Side(style="thin", color="C9A84C"),
        bottom=Side(style="thin", color="C9A84C"),
    )

    for materia in get_materias():
        df   = get_estudiantes(materia)
        ws   = wb.create_sheet(materia[:31])

        # Título
        ws.merge_cells("A1:D1")
        ws["A1"] = "✝  IGLESIA PENTECOSTAL FUENTE DE GRACIA"
        ws["A1"].fill = navy_fill; ws["A1"].font = gold_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:D2")
        ws["A2"] = f"REGISTRO ACADÉMICO — {materia.upper()}"
        ws["A2"].fill = accent_fill; ws["A2"].font = navy_font
        ws["A2"].alignment = Alignment(horizontal="center"); ws.row_dimensions[2].height = 22

        ws.merge_cells("A3:D3")
        ws["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A3"].font = Font(color="888888", italic=True, name="Calibri", size=9)
        ws["A3"].alignment = Alignment(horizontal="center"); ws.row_dimensions[3].height = 16

        for col, h in enumerate(["Nombre", "Nota", "Fecha", "Letra"], 1):
            c = ws.cell(row=4, column=col, value=h)
            c.fill = navy_fill; c.font = white_font
            c.alignment = Alignment(horizontal="center"); c.border = border
        ws.row_dimensions[4].height = 18

        for i, (_, row) in enumerate(df.iterrows()):
            r    = i + 5
            fill = PatternFill("solid", fgColor="F8F6F0" if i%2==0 else "EEE8D5")
            for col, val in enumerate([row["Nombre"], row["Nota"], row["Fecha"], row["Letra"]], 1):
                c = ws.cell(row=r, column=col, value=val)
                c.fill = fill; c.alignment = Alignment(horizontal="center"); c.border = border

        if not df.empty:
            pr = len(df) + 5
            ws.merge_cells(f"A{pr}:C{pr}")
            ws[f"A{pr}"] = "PROMEDIO DEL GRUPO"
            ws[f"A{pr}"].fill = accent_fill; ws[f"A{pr}"].font = navy_font
            ws[f"A{pr}"].alignment = Alignment(horizontal="center")
            ws[f"D{pr}"] = round(df["Nota"].mean(), 1)
            ws[f"D{pr}"].fill = accent_fill; ws[f"D{pr}"].font = navy_font
            ws[f"D{pr}"].alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 10

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return output

# ── Gráfica ───────────────────────────────────────────────────────────────────

def grafica(df, materia):
    if df.empty: return None
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11, 4))
    fig.patch.set_facecolor("#0d1b3e")
    nombres = [str(n)[:12]+"…" if len(str(n))>12 else str(n) for n in df["Nombre"]]
    notas   = df["Nota"].tolist()
    colores = ["#6ee7b7" if n>=90 else "#93c5fd" if n>=80 else
               "#fcd34d" if n>=70 else "#fdba74" if n>=60 else "#fca5a5" for n in notas]
    bars = ax1.bar(nombres, notas, color=colores, edgecolor="#c9a84c", linewidth=.7)
    ax1.set_facecolor("#0a1528"); ax1.tick_params(colors="#a0aec0", labelsize=8)
    ax1.spines[:].set_color("#1e2d4e"); ax1.set_ylim(0, 110)
    ax1.axhline(70, color="#c9a84c", linestyle="--", linewidth=.8, alpha=.6)
    ax1.set_title(f"Notas — {materia[:28]}", color="#c9a84c", fontsize=10, fontweight="bold")
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=30, ha="right")
    for bar, n in zip(bars, notas):
        ax1.text(bar.get_x()+bar.get_width()/2, bar.get_height()+1,
                 str(int(n)), ha="center", color="white", fontsize=8)
    letras = df["Letra"].value_counts()
    colmap = {"A":"#6ee7b7","B":"#93c5fd","C":"#fcd34d","D":"#fdba74","F":"#fca5a5"}
    pie_cols = [colmap.get(l,"#888") for l in letras.index]
    _, texts, autotexts = ax2.pie(
        letras.values, labels=letras.index, colors=pie_cols,
        autopct="%1.0f%%", startangle=90,
        textprops={"color":"white","fontsize":9},
        wedgeprops={"edgecolor":"#0d1b3e","linewidth":1.5}
    )
    for at in autotexts: at.set_color("#0d1b3e"); at.set_fontweight("bold")
    ax2.set_facecolor("#0a1528")
    ax2.set_title("Distribución de letras", color="#c9a84c", fontsize=10, fontweight="bold")
    plt.tight_layout(pad=2)
    return fig

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="inst-header">
  <img src="data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAC0ALQDASIAAhEBAxEB/8QAHQAAAQMFAQAAAAAAAAAAAAAAAAYHCAECBAUJA//EAFcQAAECBAQDAwQLCgoHCQAAAAECAwAEBREGBxIhCDFBE1FhFCJxgRUWGDJWkZShsbTRFyM4QlJVcoTS0yRXYmR0gpOVpcE0Q0RGsuHwKDM2N1NjdpKi/8QAHAEAAQUBAQEAAAAAAAAAAAAAAAECAwQFBwYI/8QAMREAAQMCBQMDAgYCAwAAAAAAAQACAwQRBRIhMUEGE1EUMmEHcRciU5Gx0RUWUoGh/9oADAMBAAIRAxEAPwCIkEHoEXEG2oJITewJ+iNpV1bcWAsBvzggBIvY8xYwEEcxzECEQdL7WvbnFyFFKwQEkpN7EBQ9YMW3JN+/ugQrkkC9wTcW2NuYgGmyrg3ttbvuOfqvFsECEQXud+V7mwggBIIUCUqBuCDaxhEK5QIWUq0hSdja1tvRsYt2tt64BaD1XhUIJNim50k3tfaDl3ei0XLSpJKVJIUkkEEWItzBHfFu4sd94EIg5agRvy9EAJG4JHMbH4xAALEkj0d8IhVFzZIud9gIrcgmw6Hb1QGyleanSL2Avf6YpckKKhqUbbnmP+toVCoTe1wOVthaDck87/8AW8EXbpUQU7puCDfYwIVvS8XAXBNx5u9j1i2CBCDcnkPiEEG3fBAhEBJ5E7DpB1tBAhBJJuSSe8xNnhkyLytxlkjh/EmJMLmeqs55T27/ALITTevRNOoT5qHEpFkpSNgL2ubkmITkpJ80ECw5m/TfoOt/+fOOkXBh+DVhQf0z64/FWqcQ0WT2AXR7mPJD4En+9Jz97B7mPJH4E/4rOfvYeKwgtFDuO8qSwTOe5kyP+BP+Kzn72Ke5kyP64I/xWc/fQ7FQnJSnybs5PTLMtLNJK3HnVhKEADclR2HpMR1zO4k0NuO0vAMqJl3dAqEwklBPK7aOajfkTt4GLlLS1FU60f7qCWZkW6VFQ4dcgadLqmKhhWXlGUi6nHqzNoSm3UlT1hDZ4nonCRRnFMM4fdqj6DYpkJ6ecTfl7/tQg+pUJXEGFceVijrxdmRW36ZIEgo9kVFT7qjchLTAtYkDkQmw32AJhK4KwJiHG9Xdk8LU6YmJdCgFzL4CENg8itQ825G+kXPcDa8eipMFpspfPNoN7bfus+WtkJsxqUU7N8OqSoSeUdRdHQu1qZRf1B5Vow2Z7IYrAeycmAnqUV+bJ+dwQ6UhwwiVpL07X8TOdo0ypZZkmQBsCbFa7k+pIhuMiMuJDMesVSnTtRm5Iysul1tbASbkqsQrUDceAsfGLUcGEOje9pJDd91E6SrDgDytlSHuF951IqOWlXkd/fiozTqR4nS/f4gTDk4Syw4WMVACi0uSeeP+oXV5xt0X6aFOhXxAwh8dcOOLKJLuTdCmWa6ygEqbQnsn7C52TchR8AbnoIbTCuHpar1U0ebqyaLVu07NhM60UNFd7BCl821X2F+ZsLg84nYXQVMRkp5SE8VVRE6z2qWKeGXI5VrYKB9FVnP3sVPDHkh8Cf8AFZz99DKyGM83snak1J15mZm6YTpQ3OKLrDgH/pPC5SbdN7XF0xIbKrN3C+P2vJ5R8yNUSm65F8gL/SQeSx6Nx1AjAqcOnhbnaczfIV6KqY/QixWk9zJkf8Cf8VnP3sHuYsj/AIE/4pOfvYeK8HrjL7jvKt5Qmc9zFkh8CT/es5+9iEXEphqi4QztxDh3Dsl5DTJNcuGGO1W5oCmG1q85ZUo3UpR3Jtew2sI6gRzX4xvwkcW/pyv1VmLNK4l2pTXAAJooICQeQAgttGgokQQQQIRAuxUq1wL7Am/zwbXNtvTFyElSkpSCpRNgLc77C0CELUVaTYCwAuOtup+j1R0h4L/wa8Kfrn1x+ObnM7C1zsI6R8F34NOFP1z64/FOr9oT2bp37de7lGoxbiKk4VoMxWa1NolZOXF1KJ3UTySkfjKJ2AG5MbGcmGZSWcmJhxLTLSCtalGwSkAkknuHP1RDnHeIMQ565ls4fw+FIpDDpEsDfQhANlTDlupB2HQEAbkkrh9F6l93GzRuVHUT9sWG5VMT4mxznxi80OhsuMUhshSJckpbaRyDj6he6ieQ3A5AEgkvbg3LjBOUeF38RVbs5yelWu1mKg+i5TYe9aTuE35ADcki57lxlxgqjYFw61SKOwBYBT76kjW+u261EdfDpyEMpxA1Wo5g5kUzKigOlDDbodqLwuUpVa+46hCbmx5qUBtaNE1XqX9iEZYxv9vlVu12xnfq4pPYfouIeILHjterKnpHCsi6UNtJJFk7Hs0dCsixUscrgDoBKPD1FpdBpbNLpEkzJyjKbJabTZI/5+JjywjQabhjD8pRKUwGZSVaCGxsSe8kjmoncnqSY3EZlZWGYhjdGjYKzBCGC53WFWm0O0qaZcX2aVtKSVfk3B+i8RQyer2Fss67UZ1qaqtVVMshggyrbaU6VX1DzzcGJHZu1UUbL+rTwVpWlhSUEflKGlPzkRDEX6xi1eKS0rO3GdHbroHSfTNNizXyTjQbKSJ4hMPk70mo27tKP2oMyct6Lm1g+XxLSpUUutvMB5h5aQC6CLhDtuYI67kX6i4iOMu0qYmWpdsXW6sNpAFySohIHxkROfDckin0KTkm02S0ylAHgAAPogwnEJ8+cHZR9Y4DQ4c1jIhqUwWQuM3KouayozFk23p6WKmmBNpCu1SkbtL1e+UkC4V+Mn0XOlzjyGnqDMLxTl6uZKGD2qpFpSi8yRvqZUN1D+TzHQm9gquKbA7pk2MxcPBUrWKQQ5MOs7KU0k3Ss+KCL3/JKgdrQ52UeLmMcYFkK4nSHlo7OabHJDqdlD0X39Fo9Y6rfEBUw+07ji650IQ4mN2/Ca7h6zxTiBcthfFzyWqrsiVmz5qZkjYJV+S5/wAXgdjIIEEXiNfE3lEFIdx1hRlTUyzd2oSzIsV287tkW5KSRcgcxvsQdSq4Zc0jjOimh1mYSquSDYuokBUy0LAOem5AVbqQesQ1lLFNF6mnGnI8J8ErmO7chT1iOa3GLf3SGLD/AC5Uc/5qxHSkRzW4xfwkMW/pyvX+asxn0vvVx+yaO/O8V2087qJ5W5CKKsTcDSLcr+EEaKiRpP5MEEECEG42/wA4Lm4IsCORG0Fzp02Fr928ANvXAhXECydrncne+3+UdIeC8k8NeFD/AEz64/HNwgjYi3pjo9wdPIl+GPDDzigEITOFRPIATj5irVC4ACew21Sd4w8cuU+jy2Cqc+tExUU9rOFCrEMXICDbopQ9YSR1hq8rqlmdgBqZVQsCTLrs4Qpb79OeWspA81KVAiyb726lVzGfl/JuZt8RUxWKjd2QZmDOOIO47FshLLfoJCLjqNXfEmc38QnDGA6hPMLDc0W+ylyNiFq81JHoJv6o1KmtjwymbTlgNxdyrUlFJiFQAw6k2CYioZz50yEm7Nz2DUS0s0m7jrtOeShIva5UVWAhD5eV7MqkVyo4xouFJiqTVWUS5NOSDriffqUoIKSAATa/P3oHSKjF+LPfe2es37/LHPtg9t+LSbnE9YJ8Zxz7Yw2dWUkbS1sI1XQfw0rH2cZUvfuuZ7fAdX91v/tQfdcz2+A6v7rf/ahB+2/FnwmrPy1f2we2/FnwmrPy1f2xD/s1H+gFJ+Gtb+qlpmDjXGNdy1lZfF9ORTJ2aqR0sBhbSiy2hJ1KSok2K1C3TzYbCMupVOp1NxK6lUJqcKPel90r033NtRNgbDlGJHlcQqm1UxkaLBdL6cwc4VRiFxuUp8qqaKnj+lNKbUttl3ylYSm5s2NY27yoJ9doWLubWeXaLDOBXENg+aDSnyQm+wPnc7Q19Pnp6nTPlMhOPyb2nT2jThQu3xg9Bt4RsvbdiscsS1j5Yv7Y0sIxWCiYWyR5rrzvVXSlVjNQJI32AS2ms0s8JqXdl5nAXatOoKFoVSHyFJIIIsVbixsRCGy2xnmJl5UJ3D1FoLpmJ5wPmnzEm4pSDpPnISFAi6bA3vcJHdF/tuxZ8J6x8sX9sV9t+LfhPWfli7fTG8zqylaws7IsV5N30yrCQe6l25mxnmtBSrAiilQ3HsU+R/xQ0TU3inA2NZTFrlAmKG6ZpTrbBl1sNKBN1toCt9JBtbewI7hCokcTY0np5iTlsR1hyYeWlttInF3JUQB17zD+5rYCerGRi6RNTD8/U6cx5WzMOqKnFOoBUdzc+cNSbdxjSwrqCGV5jbEA12hXmeoelpcJDTJJc7pyMJVuSxHh2QrlOcDkrOspdbPWxG4PcQbgjodo53cY34SOLf05b6qzEnOCzE5mqHVMKvuFSpJYmZdKj/q3CQoAdwUL/wBaIxcY/wCEjiz9KV+qMxDJT+mqnRrIik7kYKaMi1iesFgbm4v0G9zBBc3vfeJkqNu754IqbX5wQIVIICCDYgjvgNjawttzvzgQixvYC+3TeJz5XVs0HgOlJ9Cil1UrOsNm9iFOTzyAR4jVf1RBoKVpKQpWkm5F9ifREsDNLl+BfBUqnlNVR5CvQmamnPpQII4+5PG0+QmSvyxkpxOCSjdlh6t15SAFPzSZVCrblLaQT6rrt/Vg4rK+HKpIYebcH3pPlLwv33Sn1++hFZO56UvAGB5bDrmHZqddadcdW8h9KEqK1qVtffYED1QpZriUwvMul2YwI864ealvtlRtsLm14TGsJrKuZ+Vv8K50/i8GGysmeL5Uz+pPLUn44NSfyh8cLpzPxJx0ipIoLacNpY0KpnZMFwub+f2ltQHvdr9IUXujcI/xfuf2zf2RgSdG1jLfP2XRh9UoeI00epP5Q+ODUk9R8cO4OIzCNyfaA7/bN/ZGXN5oUTGWX2JZmn4X9iTJttNJcWtCta3ipIA0i9xYm8U6npmpp4zJJsFcoPqKysqGwMi1JTNc4L+EUEbXB8gapiul04AkPzSEK8U6gVH0BIUfVHno4zI4NG5XRqioEELpTwLqwUKuKAIo1RUm2xEsux8eUHsDXfzJU/kq/sh25riYw7ITT0mzhGafQwtTaXBMNgLCSQFAHobA+F48/dR0P4FzXylv7I9W3pSpc0ED+P7XJX/U1zHEdtNR7A138yVP5Kv7IPYGunlRKn8lX9kOx7qShfAua+Ut/ZGkxtxGU+u4UqNIkMOz1OmppgttTLc0lKmlnkoFIB2PcQYkj6QqXOAI3+39qP8AFBwGsa3PDjgecexG7X6tJPy7MkNDCHmikqcUN1AEdB86j3RJFaApopIuCCCO+IpZecRrGG8ISNGqlFqNTm5ZJSuaVNgqcuokElVySAQN78oUXuq6Sf8AdGeH62j7I0oun6qlORrb2+QvGYt1CMVnM0htfhIbKsKwFxRzFCF0Szs4/IBPe04C41t4aUer0wz/ABik+6QxYf5cr9VZheVLGUtizPqmYsk5JyRS/Pymptawo3SpKTyABuEiEFxgk+6MxWb7apX6ozGzicTmvjc/ct/hYlG4FrgNrppLEcxBzUbA/FAQNrG+2+3KC4HID1i8UQraLDwggghUI3Atfb0RcRc3FwOkX6FlCgixCBqJAAIuQncnfmQLRbpGwAUkW9N/RBZOVAkm4O59ESqmmlngjwEsAlLdYf1W3sO3nACfXYeuIspJCgoE3B23NxbrEzMNUlVU4CKb2SVKXJqmJtNvONm6g8Vf/nVvD4XZJ4z8hQzC8ZTpcNlGoVQyZoMxMUqRmH9DqXFuS6FKJDqxubHwhyPathw3vQ6b8mR9kNFwY1UTeWs3S1Hz6fUFgC/4jlnAf/spY9UPrFLEXSMqni53TqZrXRjRRxRSKS3xdv05dNlDKO0lKksllOgeYNwm3PzTvD6e1TDf5hplv6Kj7IZjHB9iuLvCs8pQCKjTuwudhcF4esklIiQI5Xh9dM8tjIcfaE2CNt3C3K0vtUw1z9gqZ8lR9kMlxQuU+myVKoVMkpWUS+6uZeSy2G9RQnSNWkC+6+vdEhibc4iXxGVMVDMuYZSrUiTZS1YG4BsVH5lCMDEqh4gIJ3XtOjKJs+JsNtG6puIczhtpAqWYwm3EakSEupwX5BSrJHzFUNnEheFCnobpVTqSrBb73Zpv1CALn0XXaMPD4884XUOr6r0+GPtzoneOFcNk3NBpl/6Kj9mKe1PDX5hpnyVH2Ru7QWj1YmeOVwUsadwtJ7VMNfmCmfJUfZCRzkw/QJTK3E0yzRqey4imPlC0S6EqCtCgLEJ2NztaHI+iG34lJ5Mjk1XiVaS80lhJ8VrCfovFilkkdMwXO4UUrGhh0Wl4bcP0abyhpExOUmSfdX2pUtxhKlH74obkg32EOT7VcN/mKm/JkfZCdyCklSOUGGmlI0qVIocI/Suq/rvC6J82Fq53undZx3SRRtyC4UO875SUb4kKVIU+WZl0JckklDKAhNy5e9htfcQzXGCf+0bisX/GlfqjMPK457cOL4dl98ZZqwSLG40yyPOsR01Iv64ZzjEJ90biwEkgLlbC/L+CMxs1rjaFh3DVVpgLvI2umjIO1xbaA7AW5233g3JtBFRWlW0EUggQvYlKgkhOkgWNr7787kn/ACEVCQTYG55A3+nuiiLGxvzNjcbfN9kegSU7AAEc78x4QqUqjd7DYi3Mg846G8LdNarHCjRKXMWU1Ny8+wvuKVTT6T8xjnqLGwtbbnHRng2B9zdha4H+2X+WPxVqXZQCOCnBoNwU0fCpV14VzWqmDqistGb1sWUbff2VKNt+pTr+KJca0nqL+mIjcXOEHqBjWWxnSw4w1UrB5bRKS3MJTbVcWIKkgHv81R6x5YByrxzjTDMrX6RmKksTAN0LmJjW2sbKQoBXNJ+PYjYiNaspIatjasvyg6HS+qzopnwuMWW6cDi2KqPOYLxrLo1rpFSuvTzKCUrsf7Mj+tD7yM0xNyjUyw4lbTqAtChyUCAQYilirILMRqgTkzNYtaqyJZlTwlCt5ZcKUk6U6ifOIBAv1MaHKbLrGOP8NKqlFxyZRph3sFy7j72popAIHmq2BSUkW9HSI30VNJTNImH5dDp5TmzSNkP5d1MudfSzKPOlQ0oSSfCIOYjqSqziGo1VRJ8rmVup3vZKiSm3gE6R6oc5GXGLsu6bVcTV/GHskyinusMsIeeVd1wJQkqC1WsASe+9oaIAAWA2jxGPBkbmxxuuuxfTmmLmyVLhrsqE2F4klkar2MrEhhzXYsYeROvp7nJh5SrW7wlI9UR5ockqpVuRp6b3mZhtuwF/fKA/zhTs4fr+aGcWLRh2vJpHkTgYSvUtIW00Q0m2g8vvd9++8WemqFk4kkebADdN+pOIGNsUDedVMjWn8oQa0flJiL3ufcy/4wk/20x+1B7n3Mv+MNP9tMftRv8Aoab9YfsVyjvy/wDBShK0flJ+OGH4xJ5x/C9DwrJqKpus1JKEJAJuE7ch/KWiEv7n3Mr+MNG3/vTH7UN1R8vcUYnzPnMJS2JxNzlIBWqoOOOKQyUqTcJuSoHUoC4tuCekXaGgpw8yd4EN12VeaeQjLk3U2KJKMUykSVOYIS1KsIZQO5KUhI+YRrMw8QsYYwXVK464j+CSylpBPvlWskeskCGA+4BmZz+6KP7eY/ahpMz6LiPDuJPajPYnfrswoI1stPOKTrURpQUqUbq96bW2uIKPCoKmX8soNtTpwiWpexli1OjwY0FyexRWsXzYK1MtmWbWQTqccUFuKv37C/6RhguMT8JHFv6ct9VZid+SeDk4Iy9p9HWEmcKe3m1J5F5dir1DZI8EiIIcYtvdI4sPcuWvt/NWYrVNQKirc9uw0CngjLIgE0fjBuRbaCAAnlDlIqkJvzPxQRSCBCyElQChe2rc+MegAIuRqJ6xRKd9J27t9vj6x6pTa5F9PXx/5wqcq6NVyLJUTcC4AEdFuDj8HHC29/8AS97W/wBsfjnaOSTtpvv4/HHRPg7sOHPC9r2/hfP+mPRUrPaErN0ucw8K07GWE53D9SRdqYR5iwPObWN0rSehBsfHkdiYitlNiyqZL5kTuF8TpcRTHntE0ACUoNrImEDqki17blPeU2iZh6Q2Ge+VknmHREuSxbla5JoV5JMKTsrr2ayN9JPXfSTex3Blw6sYwGCb2O/8+VBUxEnuM3CcaUmJedlG5mXeQ6w8nUhaFBSVJI2II5gg3iNGHpj7jnENPUeaUWMPYiUFMqOyEFSlFBv00qKknuCgTtaE7lBmlW8ra0vBuNpeZFMZd0FKgS5JXPvkj8Zs87Jvzum/IvRnNhGmZs5dIm6FMS85OS6VTFOfbUkhZIGpvV0CgLG/IgX5RP6c0chZJrG7n+FH3O8AW6ELVcVtX7DDlOpCF2VOP61AdUNi59WopiOULDM6bqrgw5Sqy++7PU+jsomO2N1pdUnUtKv5QGkE8zp33hHxz7FSDUuAOgX0N0bSenwxhO7tUtMlKe5PY6bmEAWp8s7OG4uLpQQn51A+qNzwVLU5jXETqzdSpRtSiepLhJPxmN/w20wewOKawtA1GXLCCeYsgqVb0kj4oT3BL/4wr/8AQ2v+OPYYBGG4XKfsuU9dVhnxgN4bopZbRSwivdGsr1Xp1EpUxU6nNNSsrLJK3HXFAJAHj39LdTtzimGlxsAvPEgC5SYzqxszgTAk7Vwpszq0lmRbUffvKB03HcPfHwBhH8KWEXqLg1/ElUBVVK855Qtax53ZC5QCTvc6lK/reENzIeyHEDmwmZmmnWcH0ZWzStgpJINlfy3LC/VKRbxLw5uZrYdy5pPkrXZTdV7MIlqeyoDT0Cl294kejfkB3bToHRximjF3u3/pURIHOMjtgsjPLMiSy9wut4Lbdq80kpkZYkEqVy1qHPSkkXPXYDcwzHDBgCdxPiR7MbE5dfbQ8pcqXR/pL5USpw36JNwLbEnppsdFlzgjFGdeM3MV4qdfRSAv769ukOAHZlofki5BI5b7kmJfUqQk6XTpenSEs1LSku2ltlptOlKEpFgkAcgLRLPIzD4TBGbvO58fCSNrqh+d2gCzQLbRzW4xvwj8WWBF1y1/H+CsR0pEc1uMW54kMWDc+fK2H6qzGVS+9XX6BNLclBTZOxvewv3Wv/lFvjARa+x9YgIF+oEaKiRYwRUrWLAKOw74IELLbCbpudJvubx62USoE9dze/zxYGlJ0BxKkggKFwQSDyIvzEeyEqBSgtq1KsU7EXB5W74cUpGqq2lVhcmyuY5x0U4PPwdML/rf1x+OeKAnbUCLX8b90dDuD2/udcL/AK39cfilV+0JzN07sEEEUFKkDmvljh7MKnhuotGXn2gRLzzSR2jfXSb++TffSfTsd4j/AIWwrmrlPmNTqZIuOvUeoTzbLj7aVOSjqFKGpS0c21hNzfblsoiJdx4TLKX5dxlSiErSU3SbEX228YvQYlLFGYjq08FQenYZA88KE2Pan7M40rFRCipLs2oIJ/JTdKfmSI0hPjEqmMi8CtpIMnNO9xVNuX+YiEriXJ1+TWtVBwnQKm30RM1SbaWfDYkE+sR5f/FyTylxIAK6/T9c0NLTNiaw6Cy3uTUnL0vJ4tuutomZ2Wdmi2VALKVX0mx3tYAX8IargrcQ1i7EC3FpQnyJu5J2Hn98J/MnAWbdYryak/gl6UQ3LtyrDFPmEOobabFkpBCyojmd994SUvlfmaF6GcH1pGvY+alKVfpEqAt6Y6PhuG00dE6IzAF1v+lxnFcQkq6104ZuSpe4/wA38EYRlnDNVducnE+8k5Mhx1R6XsbJHiogRF/H2Y9RzTraGK7VGKBh2XV2nYjUs7HY6Ui7jljysEjn3x6UXh+zOqKx21JlKagkefNzaOXXzUFW/ptDkYW4W2UlDuJcRrd5FTMk3oB7xrVc+sJBiWGLCqBpJkzO8qm91VOdrBIr7r03IUJnBeVVCfprBGkTCkB2cfUffL0pB849/ndLWsLKbKrh9qVWm04gzHee++q7UySnCp50ncl1y99+oBJ7yOUP7gfAGE8GSvYYfo7Esojz3yNbrh71LVdR9F7DoBCqsO6MifFmtBbTNy33PJVqKjJ1kKxJCSlafJNSUlLtS0uykIaaaQEoQByAAsAPARmc4OkHSMQkk3KvgACwVY5q8Y1/dIYs521yv1RiOlUc1uMWx4j8WW56pW/yVmLNL701+yaQgpURfcHmCDy7iPpi23MC5PgLwQbW57xoqJEEEECFnJUSU6lKV0593SNtQqDV6xK1GbpskuZYprPbzZSQOzQVBAUUkgm6lJFkgnflGnaSVAqANk8z0h6uGcgyONyQLexsrcfrrMMkcWtunDUpLVbK7H9JmfJKjhmblXTIvVAB0oTeXaALq76rXSCklN9QuARuIXWX+IuIihZYSM1g5yabwm06WZRTUpJu3ccmSgpSFpU6oqecIAIO/La0OziuWqDNcxHMLrTU5TXpDF3k8qJPQuVdTo7QFwk9oFXSRsAOW8NfLzCJXJ7JmcpEnVJ7FsvU55+jS7CUFh4pnEqcS4CdZUSlASEi1iq9torOmLxqEuW2yzMQ5k8UeHlVE1mpTMoKZLNzU2VU+RKUNOLLaFghshQKwU+aSQQQbWMarFGdfEXhedk5Kv4hdkX5yWbmpdC6bIqU40sqCVjS2QASkixsRbcCF7XqPMVTDeNHKfOzjtFxFSabMUmVmFFS5DtqkA7L26aHisbbWI3JBMafjUpzxreEqq+2w2rXO09BbWlYDTEwC0LpJAOh0XB3BBuAbwwFpOycb+VZPY04s5HEtOw7NTE6zU6klwyjJp8hpeDaSpelfZ6CUpFyNVxcbbi+DLZlcUM1jWawTL1Z9zEEqjtHpNMjT9SE6UqJ1FvSdlJOxJ3h6poMfdCvPKcKBi6rAabaw0aKCoIv4791+cRsyTNJONscGgmd9jParV/JDOhIfDfYnSXNHmhXK9ja/KBtiCbINxylS3mXxRuPiXTPzvbqqfsT2RpUmFib0FfZlJbBHmDVqI023vaPeUx1xYTdeqdCln6g5UaW2l2cYTS5H72lQJSQot6VFViUhJOqxsNjZ4/975MqP39WJQXr89ftYVfV1vy8YQ0manOT9Jw42ZuRla3hKhNSVaZukSNUaZW5KlShvZfnpIG+4tbmEu3wk18pB4PzV4nMYS1SmcNVWaqrdN0maSzTJHW3qCikaS2FKJ0KsEgnblyjZSWNuLGcpFKq0pOzDklVlobkXRI04B5SkqKQAUApuEqO4HKMHh3rkhhLBmPMQYpVU1exVapUy4mSKS66+h19QQoqIGlStlG5Niq17wsadPUSpZqZIVNSKg1iCap6X1NDR5ImWUZpQAPv+0C1EctOnxhziAbAJANEn8M414s8TSDc7QpyZnGVth0KEjTkXTrcbvZSAR57TgsQDdJ2sRekxjTi4YpZqa5mbMr5Z5DrRIU9Z8oD5l+zCUoKie1BTsLXF723hZVxEgceYTm6XpEjUMGVqpMhJ2QmZD74SP0e10+FoTeWwofsfw/9uaj7N+VzfkwT2Zliz5U9rKyTq16tGmwtbVfe0NzDwlt8rWoxtxcu4km8OtvVBVUlJZM08yKZIWS0onSoK7PSq5CgACbkKABtYY1DzF4rqzQmK1TZyemKfMTaZNl80qRQFPFwNhNlNggdoQkqICQrYkHaFbk6vFi6e+rEjaUsJquGU0hQLd1SXsrMaSrQSff9oLKsrlcWtCVzATXqtlfiWVBnaRO4fmpt9oAlDVXo788oqHcotvgHewsbbnktx4QsudxfxfydfkqDMTMyiozzLr0qx5DTVFxDQTrUCEEC2pOxIJvteEfN59cQknQafXpnFLjdNqa3W5R806RIcU0UhwABokaSpN7gc9r7w8+BZWcb4nMT16TQ0pcjh+js/fXEoSA6mTC/OUQAezQ6Bvck2FzYQzXEJShQstcNUfRpTI4mxDLJAHRMw2kEeBAB9EOjLS4AhBvbdX0TPPiNrVGq9ZpeJXZqQorSHKi8mmyIDCFkhKiC0FEHSr3oNrXNhDa5wt45XjiYn8xW3EYgn2Wpl1S0NIUtsoSltWluyU+ahIsADtuLkw83BTLyNRouPKJVN5GsCQpr4sLgP+UtpKe4hSkkeMJXjJqSa1mhS6w3bTP4dkZpNrAWWhStvCyokDg2SwCTW26afENAq+HZ9qRrUkuUmHWGphDZUlRU24kKQoFJI3SoG3MdQI1ndzI5CHT4nf8AzBpv/wAdpf1VENWdjaJ2Ou0Eph3RBAefKCHIWShW5F7A93WN3QsSVqgyVSlqVPqlWqowliaCUJKnEJWlYAUQSkhSQbpIO1r22jQAk99u8xeCT1NrbXgdqLFOHwnLfzyzSfqUpUXMTBc1KNvttKVISpCUvBAcCklspUVBtJJUCbgnmok0OdWZjtLTS04jQ1KpmhNpS1ISzakOh4PBSVJbCk/fEpNkkCwta1xDbFRJAA5+MezRsbHYKFrj6YZ22+EXTiv5x5jTlUcqL+IwqYdTLoUTJsJQAw8H2gEBASmznnGwGo7KuCRCeq2LcRVaneQVCoGYlRUHamElltJEy7YOLCgkHztO6b6Ra4AvFcCy1Imp6bFXLAQ3K62Q86lCVOdo2m27rQUdJWba08r72sdzOUDCzDc3NtVmUfQFTiGWBNAqGlD4auLAquW2lBQJSrWE2vzQNaDoEv3WVPZ1ZmT+J6diScxO49UqYHEySzKMpQ2HUlDhDYQEKKkm1ykmwG+wtrZPMzG0jjuaxrJ1hDFem2+yfmUyTJCkFKQR2ZRoAKUpFwkHa/MkkxLRsP0wGpys/JzkuKmpHkTE6lxSpfW4Ry85ICUJGolV9YII6+yMP4ZVNppya7KlSFWVNGbQhtzUlxSVaiSEpCS0FAJUoKKh6FDWgbJpuvJGaOP25tM0nE00Zj2WNZ7UtNlflfZdlr1FJJHZkp0HzdO2m0ZUnnJmZJ1up1mXxS8idqSWkzKjLMqSQ2kpbKEFBS2UgkXQEkXO+8Yc3hyiuUaZqEjVZZt9uVlnGZZdQaW664pLPapCQkEAKcXtcFPZqBBG4y3cOYdmy2puqUunoTJKJQufbU52oSSCpQWtK7kEAJKCeelO1yzPCWxSWlsQ1mVoNUoTE6U06rOMuzzRQlSnFtKUpCiogqBClqOxF+t9oU0hnHmLT6RR6XKV9DcrRVJVTh5BLksaUqSnzi2VKASpWyiQb3NzaPFugYSVNTkumsqShC1sNPuzLSRdDzKe2Cfx0KQ44oJBBPZqIJ5DClaXSWanOtpVT51SZBLrDE3UEJbDpdbSUKdQtKVkIUtQ0qHQEXBEGRp4SC4XnJ5hYwlfYwMVhQ9jJB6nSZUw0otyzuoON7pJIIWoAm5AIAIsLZtMzWzApmFJHDMjXyxSqa8l2TaRKMhbLiXe2SQ5o7T/ALzziCo35EEbQO4dws5NNkVxuXYWtvtSmYaX2KitKXWkgqClhIUpQc96QkC5JvHmih4TfQ4WayphSRLKImHm7/fglSgkhQuW7KSokWJtcJtuBrPCNfKz0Z1Zns4jmsRM4qdRUJqWRKvKEqzoDSFKUhKWyjQkpUpSgUgEFSje5JOofzKxu7gheC3a++uhrJK2FNNqUoFzWU9oU9oUlZKtOrc72jPk8J0KoyU1NSFRmFsMyUw+rUtvUwppL5SFpBurWWU2SndIXckgAq1k3Q6LK4ppUuKi1M0x6abbm3TMtlKUF1SSoqQohIUhOuxPm6rE7Akysvsg3RU8xsZ1EVQTlZLnssJTy0iWaSXhKABgXSgFOgD8W1zYquRCipWe2YTFRlXanVmKjIpqiahMyxp8qkvL7RLiwHOyKmyooAJT8RHNNzshQTiahIbfkPIJksCeMs+tLKCp5SVae1V2qAEJTcrtY3I80pJ1+KJWSZlZR5pmQlZxZWH5aTnBMNJSNOhaVBayCo9oCnUbaUnYKAJlaeEa+Vk0TH2LKLPVGYotTEkupTrM7MBthshTzLqnGinUklIStSjYEAg2II2iyv47xTXMPS+HqrU0zFNlUS6GWvJmklIYbU0yNaUhRCUKUNyb3ubkAwmTbUbbi3oi4EaCDqve4328YcWNveyS6UuJMf4sxFSvYus1NEzJ62HNHkrKDdhsst+clAVZKCU2vY8yCd4TBtc2BA6XMEEAAGgQq3gikEKheiANAPM3t80UWrfkB6oIIVO4VzZ5+EezW+kE9D9EEEBSleze6N4qN0iCCDhNcqXJB9H2RRPvvXBBC8ocqk7eqKd/ogghUBVJuQD4R5K2KfTBBCcpOFaoxbfeCCFSFXEqSHEhStIG4v76xvv3x5q3TfxgggRwrXDsn0RYBdVjysfoMEEBQFb1gSbJ12B35dIIIYUBVggggSoggggQv//Z" style="width:64px;height:64px;border-radius:50%;border:2px solid #c9a84c;object-fit:cover;flex-shrink:0"/>
  <div>
    <h1>Sistema Académico · Instituto Bíblico</h1>
    <p>Iglesia de Dios Pentecostal, M.I. · Bella Vista, Santo Domingo</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f'''<div style="text-align:center;padding:16px 0 8px">
      <img src="data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAC0ALQDASIAAhEBAxEB/8QAHQAAAQMFAQAAAAAAAAAAAAAAAAYHCAECBAUJA//EAFcQAAECBAQDAwQLCgoHCQAAAAECAwAEBREGBxIhCDFBE1FhFCJxgRUWGDJWkZShsbTRFyM4QlJVcoTS0yRXYmR0gpOVpcE0Q0RGsuHwKDM2N1NjdpKi/8QAHAEAAQUBAQEAAAAAAAAAAAAAAAECAwQFBwYI/8QAMREAAQMCBQMDAgYCAwAAAAAAAQACAwQRBRIhMUEGE1EUMmEHcRciU5Gx0RUWUoGh/9oADAMBAAIRAxEAPwCIkEHoEXEG2oJITewJ+iNpV1bcWAsBvzggBIvY8xYwEEcxzECEQdL7WvbnFyFFKwQEkpN7EBQ9YMW3JN+/ugQrkkC9wTcW2NuYgGmyrg3ttbvuOfqvFsECEQXud+V7mwggBIIUCUqBuCDaxhEK5QIWUq0hSdja1tvRsYt2tt64BaD1XhUIJNim50k3tfaDl3ei0XLSpJKVJIUkkEEWItzBHfFu4sd94EIg5agRvy9EAJG4JHMbH4xAALEkj0d8IhVFzZIud9gIrcgmw6Hb1QGyleanSL2Avf6YpckKKhqUbbnmP+toVCoTe1wOVthaDck87/8AW8EXbpUQU7puCDfYwIVvS8XAXBNx5u9j1i2CBCDcnkPiEEG3fBAhEBJ5E7DpB1tBAhBJJuSSe8xNnhkyLytxlkjh/EmJMLmeqs55T27/ALITTevRNOoT5qHEpFkpSNgL2ubkmITkpJ80ECw5m/TfoOt/+fOOkXBh+DVhQf0z64/FWqcQ0WT2AXR7mPJD4En+9Jz97B7mPJH4E/4rOfvYeKwgtFDuO8qSwTOe5kyP+BP+Kzn72Ke5kyP64I/xWc/fQ7FQnJSnybs5PTLMtLNJK3HnVhKEADclR2HpMR1zO4k0NuO0vAMqJl3dAqEwklBPK7aOajfkTt4GLlLS1FU60f7qCWZkW6VFQ4dcgadLqmKhhWXlGUi6nHqzNoSm3UlT1hDZ4nonCRRnFMM4fdqj6DYpkJ6ecTfl7/tQg+pUJXEGFceVijrxdmRW36ZIEgo9kVFT7qjchLTAtYkDkQmw32AJhK4KwJiHG9Xdk8LU6YmJdCgFzL4CENg8itQ825G+kXPcDa8eipMFpspfPNoN7bfus+WtkJsxqUU7N8OqSoSeUdRdHQu1qZRf1B5Vow2Z7IYrAeycmAnqUV+bJ+dwQ6UhwwiVpL07X8TOdo0ypZZkmQBsCbFa7k+pIhuMiMuJDMesVSnTtRm5Iysul1tbASbkqsQrUDceAsfGLUcGEOje9pJDd91E6SrDgDytlSHuF951IqOWlXkd/fiozTqR4nS/f4gTDk4Syw4WMVACi0uSeeP+oXV5xt0X6aFOhXxAwh8dcOOLKJLuTdCmWa6ygEqbQnsn7C52TchR8AbnoIbTCuHpar1U0ebqyaLVu07NhM60UNFd7BCl821X2F+ZsLg84nYXQVMRkp5SE8VVRE6z2qWKeGXI5VrYKB9FVnP3sVPDHkh8Cf8AFZz99DKyGM83snak1J15mZm6YTpQ3OKLrDgH/pPC5SbdN7XF0xIbKrN3C+P2vJ5R8yNUSm65F8gL/SQeSx6Nx1AjAqcOnhbnaczfIV6KqY/QixWk9zJkf8Cf8VnP3sHuYsj/AIE/4pOfvYeK8HrjL7jvKt5Qmc9zFkh8CT/es5+9iEXEphqi4QztxDh3Dsl5DTJNcuGGO1W5oCmG1q85ZUo3UpR3Jtew2sI6gRzX4xvwkcW/pyv1VmLNK4l2pTXAAJooICQeQAgttGgokQQQQIRAuxUq1wL7Am/zwbXNtvTFyElSkpSCpRNgLc77C0CELUVaTYCwAuOtup+j1R0h4L/wa8Kfrn1x+ObnM7C1zsI6R8F34NOFP1z64/FOr9oT2bp37de7lGoxbiKk4VoMxWa1NolZOXF1KJ3UTySkfjKJ2AG5MbGcmGZSWcmJhxLTLSCtalGwSkAkknuHP1RDnHeIMQ565ls4fw+FIpDDpEsDfQhANlTDlupB2HQEAbkkrh9F6l93GzRuVHUT9sWG5VMT4mxznxi80OhsuMUhshSJckpbaRyDj6he6ieQ3A5AEgkvbg3LjBOUeF38RVbs5yelWu1mKg+i5TYe9aTuE35ADcki57lxlxgqjYFw61SKOwBYBT76kjW+u261EdfDpyEMpxA1Wo5g5kUzKigOlDDbodqLwuUpVa+46hCbmx5qUBtaNE1XqX9iEZYxv9vlVu12xnfq4pPYfouIeILHjterKnpHCsi6UNtJJFk7Hs0dCsixUscrgDoBKPD1FpdBpbNLpEkzJyjKbJabTZI/5+JjywjQabhjD8pRKUwGZSVaCGxsSe8kjmoncnqSY3EZlZWGYhjdGjYKzBCGC53WFWm0O0qaZcX2aVtKSVfk3B+i8RQyer2Fss67UZ1qaqtVVMshggyrbaU6VX1DzzcGJHZu1UUbL+rTwVpWlhSUEflKGlPzkRDEX6xi1eKS0rO3GdHbroHSfTNNizXyTjQbKSJ4hMPk70mo27tKP2oMyct6Lm1g+XxLSpUUutvMB5h5aQC6CLhDtuYI67kX6i4iOMu0qYmWpdsXW6sNpAFySohIHxkROfDckin0KTkm02S0ylAHgAAPogwnEJ8+cHZR9Y4DQ4c1jIhqUwWQuM3KouayozFk23p6WKmmBNpCu1SkbtL1e+UkC4V+Mn0XOlzjyGnqDMLxTl6uZKGD2qpFpSi8yRvqZUN1D+TzHQm9gquKbA7pk2MxcPBUrWKQQ5MOs7KU0k3Ss+KCL3/JKgdrQ52UeLmMcYFkK4nSHlo7OabHJDqdlD0X39Fo9Y6rfEBUw+07ji650IQ4mN2/Ca7h6zxTiBcthfFzyWqrsiVmz5qZkjYJV+S5/wAXgdjIIEEXiNfE3lEFIdx1hRlTUyzd2oSzIsV287tkW5KSRcgcxvsQdSq4Zc0jjOimh1mYSquSDYuokBUy0LAOem5AVbqQesQ1lLFNF6mnGnI8J8ErmO7chT1iOa3GLf3SGLD/AC5Uc/5qxHSkRzW4xfwkMW/pyvX+asxn0vvVx+yaO/O8V2087qJ5W5CKKsTcDSLcr+EEaKiRpP5MEEECEG42/wA4Lm4IsCORG0Fzp02Fr928ANvXAhXECydrncne+3+UdIeC8k8NeFD/AEz64/HNwgjYi3pjo9wdPIl+GPDDzigEITOFRPIATj5irVC4ACew21Sd4w8cuU+jy2Cqc+tExUU9rOFCrEMXICDbopQ9YSR1hq8rqlmdgBqZVQsCTLrs4Qpb79OeWspA81KVAiyb726lVzGfl/JuZt8RUxWKjd2QZmDOOIO47FshLLfoJCLjqNXfEmc38QnDGA6hPMLDc0W+ylyNiFq81JHoJv6o1KmtjwymbTlgNxdyrUlFJiFQAw6k2CYioZz50yEm7Nz2DUS0s0m7jrtOeShIva5UVWAhD5eV7MqkVyo4xouFJiqTVWUS5NOSDriffqUoIKSAATa/P3oHSKjF+LPfe2es37/LHPtg9t+LSbnE9YJ8Zxz7Yw2dWUkbS1sI1XQfw0rH2cZUvfuuZ7fAdX91v/tQfdcz2+A6v7rf/ahB+2/FnwmrPy1f2we2/FnwmrPy1f2xD/s1H+gFJ+Gtb+qlpmDjXGNdy1lZfF9ORTJ2aqR0sBhbSiy2hJ1KSok2K1C3TzYbCMupVOp1NxK6lUJqcKPel90r033NtRNgbDlGJHlcQqm1UxkaLBdL6cwc4VRiFxuUp8qqaKnj+lNKbUttl3ylYSm5s2NY27yoJ9doWLubWeXaLDOBXENg+aDSnyQm+wPnc7Q19Pnp6nTPlMhOPyb2nT2jThQu3xg9Bt4RsvbdiscsS1j5Yv7Y0sIxWCiYWyR5rrzvVXSlVjNQJI32AS2ms0s8JqXdl5nAXatOoKFoVSHyFJIIIsVbixsRCGy2xnmJl5UJ3D1FoLpmJ5wPmnzEm4pSDpPnISFAi6bA3vcJHdF/tuxZ8J6x8sX9sV9t+LfhPWfli7fTG8zqylaws7IsV5N30yrCQe6l25mxnmtBSrAiilQ3HsU+R/xQ0TU3inA2NZTFrlAmKG6ZpTrbBl1sNKBN1toCt9JBtbewI7hCokcTY0np5iTlsR1hyYeWlttInF3JUQB17zD+5rYCerGRi6RNTD8/U6cx5WzMOqKnFOoBUdzc+cNSbdxjSwrqCGV5jbEA12hXmeoelpcJDTJJc7pyMJVuSxHh2QrlOcDkrOspdbPWxG4PcQbgjodo53cY34SOLf05b6qzEnOCzE5mqHVMKvuFSpJYmZdKj/q3CQoAdwUL/wBaIxcY/wCEjiz9KV+qMxDJT+mqnRrIik7kYKaMi1iesFgbm4v0G9zBBc3vfeJkqNu754IqbX5wQIVIICCDYgjvgNjawttzvzgQixvYC+3TeJz5XVs0HgOlJ9Cil1UrOsNm9iFOTzyAR4jVf1RBoKVpKQpWkm5F9ifREsDNLl+BfBUqnlNVR5CvQmamnPpQII4+5PG0+QmSvyxkpxOCSjdlh6t15SAFPzSZVCrblLaQT6rrt/Vg4rK+HKpIYebcH3pPlLwv33Sn1++hFZO56UvAGB5bDrmHZqddadcdW8h9KEqK1qVtffYED1QpZriUwvMul2YwI864ealvtlRtsLm14TGsJrKuZ+Vv8K50/i8GGysmeL5Uz+pPLUn44NSfyh8cLpzPxJx0ipIoLacNpY0KpnZMFwub+f2ltQHvdr9IUXujcI/xfuf2zf2RgSdG1jLfP2XRh9UoeI00epP5Q+ODUk9R8cO4OIzCNyfaA7/bN/ZGXN5oUTGWX2JZmn4X9iTJttNJcWtCta3ipIA0i9xYm8U6npmpp4zJJsFcoPqKysqGwMi1JTNc4L+EUEbXB8gapiul04AkPzSEK8U6gVH0BIUfVHno4zI4NG5XRqioEELpTwLqwUKuKAIo1RUm2xEsux8eUHsDXfzJU/kq/sh25riYw7ITT0mzhGafQwtTaXBMNgLCSQFAHobA+F48/dR0P4FzXylv7I9W3pSpc0ED+P7XJX/U1zHEdtNR7A138yVP5Kv7IPYGunlRKn8lX9kOx7qShfAua+Ut/ZGkxtxGU+u4UqNIkMOz1OmppgttTLc0lKmlnkoFIB2PcQYkj6QqXOAI3+39qP8AFBwGsa3PDjgecexG7X6tJPy7MkNDCHmikqcUN1AEdB86j3RJFaApopIuCCCO+IpZecRrGG8ISNGqlFqNTm5ZJSuaVNgqcuokElVySAQN78oUXuq6Sf8AdGeH62j7I0oun6qlORrb2+QvGYt1CMVnM0htfhIbKsKwFxRzFCF0Szs4/IBPe04C41t4aUer0wz/ABik+6QxYf5cr9VZheVLGUtizPqmYsk5JyRS/Pymptawo3SpKTyABuEiEFxgk+6MxWb7apX6ozGzicTmvjc/ct/hYlG4FrgNrppLEcxBzUbA/FAQNrG+2+3KC4HID1i8UQraLDwggghUI3Atfb0RcRc3FwOkX6FlCgixCBqJAAIuQncnfmQLRbpGwAUkW9N/RBZOVAkm4O59ESqmmlngjwEsAlLdYf1W3sO3nACfXYeuIspJCgoE3B23NxbrEzMNUlVU4CKb2SVKXJqmJtNvONm6g8Vf/nVvD4XZJ4z8hQzC8ZTpcNlGoVQyZoMxMUqRmH9DqXFuS6FKJDqxubHwhyPathw3vQ6b8mR9kNFwY1UTeWs3S1Hz6fUFgC/4jlnAf/spY9UPrFLEXSMqni53TqZrXRjRRxRSKS3xdv05dNlDKO0lKksllOgeYNwm3PzTvD6e1TDf5hplv6Kj7IZjHB9iuLvCs8pQCKjTuwudhcF4esklIiQI5Xh9dM8tjIcfaE2CNt3C3K0vtUw1z9gqZ8lR9kMlxQuU+myVKoVMkpWUS+6uZeSy2G9RQnSNWkC+6+vdEhibc4iXxGVMVDMuYZSrUiTZS1YG4BsVH5lCMDEqh4gIJ3XtOjKJs+JsNtG6puIczhtpAqWYwm3EakSEupwX5BSrJHzFUNnEheFCnobpVTqSrBb73Zpv1CALn0XXaMPD4884XUOr6r0+GPtzoneOFcNk3NBpl/6Kj9mKe1PDX5hpnyVH2Ru7QWj1YmeOVwUsadwtJ7VMNfmCmfJUfZCRzkw/QJTK3E0yzRqey4imPlC0S6EqCtCgLEJ2NztaHI+iG34lJ5Mjk1XiVaS80lhJ8VrCfovFilkkdMwXO4UUrGhh0Wl4bcP0abyhpExOUmSfdX2pUtxhKlH74obkg32EOT7VcN/mKm/JkfZCdyCklSOUGGmlI0qVIocI/Suq/rvC6J82Fq53undZx3SRRtyC4UO875SUb4kKVIU+WZl0JckklDKAhNy5e9htfcQzXGCf+0bisX/GlfqjMPK457cOL4dl98ZZqwSLG40yyPOsR01Iv64ZzjEJ90biwEkgLlbC/L+CMxs1rjaFh3DVVpgLvI2umjIO1xbaA7AW5233g3JtBFRWlW0EUggQvYlKgkhOkgWNr7787kn/ACEVCQTYG55A3+nuiiLGxvzNjcbfN9kegSU7AAEc78x4QqUqjd7DYi3Mg846G8LdNarHCjRKXMWU1Ny8+wvuKVTT6T8xjnqLGwtbbnHRng2B9zdha4H+2X+WPxVqXZQCOCnBoNwU0fCpV14VzWqmDqistGb1sWUbff2VKNt+pTr+KJca0nqL+mIjcXOEHqBjWWxnSw4w1UrB5bRKS3MJTbVcWIKkgHv81R6x5YByrxzjTDMrX6RmKksTAN0LmJjW2sbKQoBXNJ+PYjYiNaspIatjasvyg6HS+qzopnwuMWW6cDi2KqPOYLxrLo1rpFSuvTzKCUrsf7Mj+tD7yM0xNyjUyw4lbTqAtChyUCAQYilirILMRqgTkzNYtaqyJZlTwlCt5ZcKUk6U6ifOIBAv1MaHKbLrGOP8NKqlFxyZRph3sFy7j72popAIHmq2BSUkW9HSI30VNJTNImH5dDp5TmzSNkP5d1MudfSzKPOlQ0oSSfCIOYjqSqziGo1VRJ8rmVup3vZKiSm3gE6R6oc5GXGLsu6bVcTV/GHskyinusMsIeeVd1wJQkqC1WsASe+9oaIAAWA2jxGPBkbmxxuuuxfTmmLmyVLhrsqE2F4klkar2MrEhhzXYsYeROvp7nJh5SrW7wlI9UR5ockqpVuRp6b3mZhtuwF/fKA/zhTs4fr+aGcWLRh2vJpHkTgYSvUtIW00Q0m2g8vvd9++8WemqFk4kkebADdN+pOIGNsUDedVMjWn8oQa0flJiL3ufcy/4wk/20x+1B7n3Mv+MNP9tMftRv8Aoab9YfsVyjvy/wDBShK0flJ+OGH4xJ5x/C9DwrJqKpus1JKEJAJuE7ch/KWiEv7n3Mr+MNG3/vTH7UN1R8vcUYnzPnMJS2JxNzlIBWqoOOOKQyUqTcJuSoHUoC4tuCekXaGgpw8yd4EN12VeaeQjLk3U2KJKMUykSVOYIS1KsIZQO5KUhI+YRrMw8QsYYwXVK464j+CSylpBPvlWskeskCGA+4BmZz+6KP7eY/ahpMz6LiPDuJPajPYnfrswoI1stPOKTrURpQUqUbq96bW2uIKPCoKmX8soNtTpwiWpexli1OjwY0FyexRWsXzYK1MtmWbWQTqccUFuKv37C/6RhguMT8JHFv6ct9VZid+SeDk4Iy9p9HWEmcKe3m1J5F5dir1DZI8EiIIcYtvdI4sPcuWvt/NWYrVNQKirc9uw0CngjLIgE0fjBuRbaCAAnlDlIqkJvzPxQRSCBCyElQChe2rc+MegAIuRqJ6xRKd9J27t9vj6x6pTa5F9PXx/5wqcq6NVyLJUTcC4AEdFuDj8HHC29/8AS97W/wBsfjnaOSTtpvv4/HHRPg7sOHPC9r2/hfP+mPRUrPaErN0ucw8K07GWE53D9SRdqYR5iwPObWN0rSehBsfHkdiYitlNiyqZL5kTuF8TpcRTHntE0ACUoNrImEDqki17blPeU2iZh6Q2Ge+VknmHREuSxbla5JoV5JMKTsrr2ayN9JPXfSTex3Blw6sYwGCb2O/8+VBUxEnuM3CcaUmJedlG5mXeQ6w8nUhaFBSVJI2II5gg3iNGHpj7jnENPUeaUWMPYiUFMqOyEFSlFBv00qKknuCgTtaE7lBmlW8ra0vBuNpeZFMZd0FKgS5JXPvkj8Zs87Jvzum/IvRnNhGmZs5dIm6FMS85OS6VTFOfbUkhZIGpvV0CgLG/IgX5RP6c0chZJrG7n+FH3O8AW6ELVcVtX7DDlOpCF2VOP61AdUNi59WopiOULDM6bqrgw5Sqy++7PU+jsomO2N1pdUnUtKv5QGkE8zp33hHxz7FSDUuAOgX0N0bSenwxhO7tUtMlKe5PY6bmEAWp8s7OG4uLpQQn51A+qNzwVLU5jXETqzdSpRtSiepLhJPxmN/w20wewOKawtA1GXLCCeYsgqVb0kj4oT3BL/4wr/8AQ2v+OPYYBGG4XKfsuU9dVhnxgN4bopZbRSwivdGsr1Xp1EpUxU6nNNSsrLJK3HXFAJAHj39LdTtzimGlxsAvPEgC5SYzqxszgTAk7Vwpszq0lmRbUffvKB03HcPfHwBhH8KWEXqLg1/ElUBVVK855Qtax53ZC5QCTvc6lK/reENzIeyHEDmwmZmmnWcH0ZWzStgpJINlfy3LC/VKRbxLw5uZrYdy5pPkrXZTdV7MIlqeyoDT0Cl294kejfkB3bToHRximjF3u3/pURIHOMjtgsjPLMiSy9wut4Lbdq80kpkZYkEqVy1qHPSkkXPXYDcwzHDBgCdxPiR7MbE5dfbQ8pcqXR/pL5USpw36JNwLbEnppsdFlzgjFGdeM3MV4qdfRSAv769ukOAHZlofki5BI5b7kmJfUqQk6XTpenSEs1LSku2ltlptOlKEpFgkAcgLRLPIzD4TBGbvO58fCSNrqh+d2gCzQLbRzW4xvwj8WWBF1y1/H+CsR0pEc1uMW54kMWDc+fK2H6qzGVS+9XX6BNLclBTZOxvewv3Wv/lFvjARa+x9YgIF+oEaKiRYwRUrWLAKOw74IELLbCbpudJvubx62USoE9dze/zxYGlJ0BxKkggKFwQSDyIvzEeyEqBSgtq1KsU7EXB5W74cUpGqq2lVhcmyuY5x0U4PPwdML/rf1x+OeKAnbUCLX8b90dDuD2/udcL/AK39cfilV+0JzN07sEEEUFKkDmvljh7MKnhuotGXn2gRLzzSR2jfXSb++TffSfTsd4j/AIWwrmrlPmNTqZIuOvUeoTzbLj7aVOSjqFKGpS0c21hNzfblsoiJdx4TLKX5dxlSiErSU3SbEX228YvQYlLFGYjq08FQenYZA88KE2Pan7M40rFRCipLs2oIJ/JTdKfmSI0hPjEqmMi8CtpIMnNO9xVNuX+YiEriXJ1+TWtVBwnQKm30RM1SbaWfDYkE+sR5f/FyTylxIAK6/T9c0NLTNiaw6Cy3uTUnL0vJ4tuutomZ2Wdmi2VALKVX0mx3tYAX8IargrcQ1i7EC3FpQnyJu5J2Hn98J/MnAWbdYryak/gl6UQ3LtyrDFPmEOobabFkpBCyojmd994SUvlfmaF6GcH1pGvY+alKVfpEqAt6Y6PhuG00dE6IzAF1v+lxnFcQkq6104ZuSpe4/wA38EYRlnDNVducnE+8k5Mhx1R6XsbJHiogRF/H2Y9RzTraGK7VGKBh2XV2nYjUs7HY6Ui7jljysEjn3x6UXh+zOqKx21JlKagkefNzaOXXzUFW/ptDkYW4W2UlDuJcRrd5FTMk3oB7xrVc+sJBiWGLCqBpJkzO8qm91VOdrBIr7r03IUJnBeVVCfprBGkTCkB2cfUffL0pB849/ndLWsLKbKrh9qVWm04gzHee++q7UySnCp50ncl1y99+oBJ7yOUP7gfAGE8GSvYYfo7Esojz3yNbrh71LVdR9F7DoBCqsO6MifFmtBbTNy33PJVqKjJ1kKxJCSlafJNSUlLtS0uykIaaaQEoQByAAsAPARmc4OkHSMQkk3KvgACwVY5q8Y1/dIYs521yv1RiOlUc1uMWx4j8WW56pW/yVmLNL701+yaQgpURfcHmCDy7iPpi23MC5PgLwQbW57xoqJEEEECFnJUSU6lKV0593SNtQqDV6xK1GbpskuZYprPbzZSQOzQVBAUUkgm6lJFkgnflGnaSVAqANk8z0h6uGcgyONyQLexsrcfrrMMkcWtunDUpLVbK7H9JmfJKjhmblXTIvVAB0oTeXaALq76rXSCklN9QuARuIXWX+IuIihZYSM1g5yabwm06WZRTUpJu3ccmSgpSFpU6oqecIAIO/La0OziuWqDNcxHMLrTU5TXpDF3k8qJPQuVdTo7QFwk9oFXSRsAOW8NfLzCJXJ7JmcpEnVJ7FsvU55+jS7CUFh4pnEqcS4CdZUSlASEi1iq9torOmLxqEuW2yzMQ5k8UeHlVE1mpTMoKZLNzU2VU+RKUNOLLaFghshQKwU+aSQQQbWMarFGdfEXhedk5Kv4hdkX5yWbmpdC6bIqU40sqCVjS2QASkixsRbcCF7XqPMVTDeNHKfOzjtFxFSabMUmVmFFS5DtqkA7L26aHisbbWI3JBMafjUpzxreEqq+2w2rXO09BbWlYDTEwC0LpJAOh0XB3BBuAbwwFpOycb+VZPY04s5HEtOw7NTE6zU6klwyjJp8hpeDaSpelfZ6CUpFyNVxcbbi+DLZlcUM1jWawTL1Z9zEEqjtHpNMjT9SE6UqJ1FvSdlJOxJ3h6poMfdCvPKcKBi6rAabaw0aKCoIv4791+cRsyTNJONscGgmd9jParV/JDOhIfDfYnSXNHmhXK9ja/KBtiCbINxylS3mXxRuPiXTPzvbqqfsT2RpUmFib0FfZlJbBHmDVqI023vaPeUx1xYTdeqdCln6g5UaW2l2cYTS5H72lQJSQot6VFViUhJOqxsNjZ4/975MqP39WJQXr89ftYVfV1vy8YQ0manOT9Jw42ZuRla3hKhNSVaZukSNUaZW5KlShvZfnpIG+4tbmEu3wk18pB4PzV4nMYS1SmcNVWaqrdN0maSzTJHW3qCikaS2FKJ0KsEgnblyjZSWNuLGcpFKq0pOzDklVlobkXRI04B5SkqKQAUApuEqO4HKMHh3rkhhLBmPMQYpVU1exVapUy4mSKS66+h19QQoqIGlStlG5Niq17wsadPUSpZqZIVNSKg1iCap6X1NDR5ImWUZpQAPv+0C1EctOnxhziAbAJANEn8M414s8TSDc7QpyZnGVth0KEjTkXTrcbvZSAR57TgsQDdJ2sRekxjTi4YpZqa5mbMr5Z5DrRIU9Z8oD5l+zCUoKie1BTsLXF723hZVxEgceYTm6XpEjUMGVqpMhJ2QmZD74SP0e10+FoTeWwofsfw/9uaj7N+VzfkwT2Zliz5U9rKyTq16tGmwtbVfe0NzDwlt8rWoxtxcu4km8OtvVBVUlJZM08yKZIWS0onSoK7PSq5CgACbkKABtYY1DzF4rqzQmK1TZyemKfMTaZNl80qRQFPFwNhNlNggdoQkqICQrYkHaFbk6vFi6e+rEjaUsJquGU0hQLd1SXsrMaSrQSff9oLKsrlcWtCVzATXqtlfiWVBnaRO4fmpt9oAlDVXo788oqHcotvgHewsbbnktx4QsudxfxfydfkqDMTMyiozzLr0qx5DTVFxDQTrUCEEC2pOxIJvteEfN59cQknQafXpnFLjdNqa3W5R806RIcU0UhwABokaSpN7gc9r7w8+BZWcb4nMT16TQ0pcjh+js/fXEoSA6mTC/OUQAezQ6Bvck2FzYQzXEJShQstcNUfRpTI4mxDLJAHRMw2kEeBAB9EOjLS4AhBvbdX0TPPiNrVGq9ZpeJXZqQorSHKi8mmyIDCFkhKiC0FEHSr3oNrXNhDa5wt45XjiYn8xW3EYgn2Wpl1S0NIUtsoSltWluyU+ahIsADtuLkw83BTLyNRouPKJVN5GsCQpr4sLgP+UtpKe4hSkkeMJXjJqSa1mhS6w3bTP4dkZpNrAWWhStvCyokDg2SwCTW26afENAq+HZ9qRrUkuUmHWGphDZUlRU24kKQoFJI3SoG3MdQI1ndzI5CHT4nf8AzBpv/wAdpf1VENWdjaJ2Ou0Eph3RBAefKCHIWShW5F7A93WN3QsSVqgyVSlqVPqlWqowliaCUJKnEJWlYAUQSkhSQbpIO1r22jQAk99u8xeCT1NrbXgdqLFOHwnLfzyzSfqUpUXMTBc1KNvttKVISpCUvBAcCklspUVBtJJUCbgnmok0OdWZjtLTS04jQ1KpmhNpS1ISzakOh4PBSVJbCk/fEpNkkCwta1xDbFRJAA5+MezRsbHYKFrj6YZ22+EXTiv5x5jTlUcqL+IwqYdTLoUTJsJQAw8H2gEBASmznnGwGo7KuCRCeq2LcRVaneQVCoGYlRUHamElltJEy7YOLCgkHztO6b6Ra4AvFcCy1Imp6bFXLAQ3K62Q86lCVOdo2m27rQUdJWba08r72sdzOUDCzDc3NtVmUfQFTiGWBNAqGlD4auLAquW2lBQJSrWE2vzQNaDoEv3WVPZ1ZmT+J6diScxO49UqYHEySzKMpQ2HUlDhDYQEKKkm1ykmwG+wtrZPMzG0jjuaxrJ1hDFem2+yfmUyTJCkFKQR2ZRoAKUpFwkHa/MkkxLRsP0wGpys/JzkuKmpHkTE6lxSpfW4Ry85ICUJGolV9YII6+yMP4ZVNppya7KlSFWVNGbQhtzUlxSVaiSEpCS0FAJUoKKh6FDWgbJpuvJGaOP25tM0nE00Zj2WNZ7UtNlflfZdlr1FJJHZkp0HzdO2m0ZUnnJmZJ1up1mXxS8idqSWkzKjLMqSQ2kpbKEFBS2UgkXQEkXO+8Yc3hyiuUaZqEjVZZt9uVlnGZZdQaW664pLPapCQkEAKcXtcFPZqBBG4y3cOYdmy2puqUunoTJKJQufbU52oSSCpQWtK7kEAJKCeelO1yzPCWxSWlsQ1mVoNUoTE6U06rOMuzzRQlSnFtKUpCiogqBClqOxF+t9oU0hnHmLT6RR6XKV9DcrRVJVTh5BLksaUqSnzi2VKASpWyiQb3NzaPFugYSVNTkumsqShC1sNPuzLSRdDzKe2Cfx0KQ44oJBBPZqIJ5DClaXSWanOtpVT51SZBLrDE3UEJbDpdbSUKdQtKVkIUtQ0qHQEXBEGRp4SC4XnJ5hYwlfYwMVhQ9jJB6nSZUw0otyzuoON7pJIIWoAm5AIAIsLZtMzWzApmFJHDMjXyxSqa8l2TaRKMhbLiXe2SQ5o7T/ALzziCo35EEbQO4dws5NNkVxuXYWtvtSmYaX2KitKXWkgqClhIUpQc96QkC5JvHmih4TfQ4WayphSRLKImHm7/fglSgkhQuW7KSokWJtcJtuBrPCNfKz0Z1Zns4jmsRM4qdRUJqWRKvKEqzoDSFKUhKWyjQkpUpSgUgEFSje5JOofzKxu7gheC3a++uhrJK2FNNqUoFzWU9oU9oUlZKtOrc72jPk8J0KoyU1NSFRmFsMyUw+rUtvUwppL5SFpBurWWU2SndIXckgAq1k3Q6LK4ppUuKi1M0x6abbm3TMtlKUF1SSoqQohIUhOuxPm6rE7Akysvsg3RU8xsZ1EVQTlZLnssJTy0iWaSXhKABgXSgFOgD8W1zYquRCipWe2YTFRlXanVmKjIpqiahMyxp8qkvL7RLiwHOyKmyooAJT8RHNNzshQTiahIbfkPIJksCeMs+tLKCp5SVae1V2qAEJTcrtY3I80pJ1+KJWSZlZR5pmQlZxZWH5aTnBMNJSNOhaVBayCo9oCnUbaUnYKAJlaeEa+Vk0TH2LKLPVGYotTEkupTrM7MBthshTzLqnGinUklIStSjYEAg2II2iyv47xTXMPS+HqrU0zFNlUS6GWvJmklIYbU0yNaUhRCUKUNyb3ubkAwmTbUbbi3oi4EaCDqve4328YcWNveyS6UuJMf4sxFSvYus1NEzJ62HNHkrKDdhsst+clAVZKCU2vY8yCd4TBtc2BA6XMEEAAGgQq3gikEKheiANAPM3t80UWrfkB6oIIVO4VzZ5+EezW+kE9D9EEEBSleze6N4qN0iCCDhNcqXJB9H2RRPvvXBBC8ocqk7eqKd/ogghUBVJuQD4R5K2KfTBBCcpOFaoxbfeCCFSFXEqSHEhStIG4v76xvv3x5q3TfxgggRwrXDsn0RYBdVjysfoMEEBQFb1gSbJ12B35dIIIYUBVggggSoggggQv//Z" style="width:72px;height:72px;border-radius:50%;border:2px solid #c9a84c"/>
      <div style="color:#e8c97a;font-size:.75rem;letter-spacing:1px;margin-top:8px;text-transform:uppercase">Instituto Bíblico</div>
    </div>''', unsafe_allow_html=True)
    st.markdown("### ✝ Navegación")
    st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)
    pagina = st.radio("", [
        "📊 Dashboard",
        "📖 Ver Materia",
        "➕ Crear Materia",
        "📈 Estadísticas",
        "📥 Exportar Excel"
    ], label_visibility="collapsed")
    st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)
    try:
        materias  = get_materias()
        total_est = sum(len(get_estudiantes(m)) for m in materias)
        st.markdown(f"**{len(materias)}** materias · **{total_est}** estudiantes")
    except:
        st.markdown("Conectando…")
    st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)
    # Versículo del día — rota automáticamente cada día
    versiculos = [
        ("Instruye al sabio, y se hará más sabio.", "Proverbios 9:9"),
        ("Todo lo puedo en Cristo que me fortalece.", "Filipenses 4:13"),
        ("El Señor es mi pastor, nada me faltará.", "Salmos 23:1"),
        ("Porque de tal manera amó Dios al mundo...", "Juan 3:16"),
        ("Confía en el Señor con todo tu corazón.", "Proverbios 3:5"),
        ("Busca primero el reino de Dios.", "Mateo 6:33"),
        ("El temor del Señor es el principio de la sabiduría.", "Proverbios 9:10"),
        ("Lámpara es tu palabra a mis pies.", "Salmos 119:105"),
        ("Jehová es mi luz y mi salvación.", "Salmos 27:1"),
        ("Esfuérzate y sé valiente.", "Josué 1:9"),
        ("El justo vivirá por la fe.", "Romanos 1:17"),
        ("Dios es amor.", "1 Juan 4:8"),
        ("Yo soy el camino, la verdad y la vida.", "Juan 14:6"),
        ("Fiel es el que os llama.", "1 Tesalonicenses 5:24"),
        ("El Señor tu Dios está contigo.", "Sofonías 3:17"),
        ("Mas los que esperan en Jehová tendrán nuevas fuerzas.", "Isaías 40:31"),
        ("Amarás al Señor tu Dios con todo tu corazón.", "Mateo 22:37"),
        ("La gracia del Señor sea con todos vosotros.", "Apocalipsis 22:21"),
        ("Bienaventurados los de limpio corazón.", "Mateo 5:8"),
        ("Encomienda al Señor tu camino.", "Salmos 37:5"),
        ("Sed fuertes y valientes. No temáis.", "Deuteronomio 31:6"),
        ("Tu fe te ha salvado.", "Lucas 7:50"),
        ("El amor nunca deja de ser.", "1 Corintios 13:8"),
        ("Den gracias en todo.", "1 Tesalonicenses 5:18"),
        ("Conocerán la verdad y la verdad los hará libres.", "Juan 8:32"),
        ("No os afanéis por nada.", "Filipenses 4:6"),
        ("Yo estoy con vosotros todos los días.", "Mateo 28:20"),
        ("El que comenzó en vosotros la buena obra, la perfeccionará.", "Filipenses 1:6"),
        ("Porque yo sé los planes que tengo para vosotros.", "Jeremías 29:11"),
        ("Gracia y paz a vosotros de parte de Dios.", "Romanos 1:7"),
        ("El Señor es bueno; para siempre es su misericordia.", "Salmos 100:5"),
        ("Pedid y se os dará.", "Mateo 7:7"),
        ("Jesús dijo: Yo soy la resurrección y la vida.", "Juan 11:25"),
        ("Ama a tu prójimo como a ti mismo.", "Mateo 22:39"),
        ("El Señor bendiga tu salida y tu entrada.", "Salmos 121:8"),
        ("La fe es la certeza de lo que se espera.", "Hebreos 11:1"),
        ("Dios nos dio espíritu de poder, amor y dominio propio.", "2 Timoteo 1:7"),
    ]
    from datetime import date
    idx = date.today().timetuple().tm_yday % len(versiculos)
    texto, referencia = versiculos[idx]
    st.markdown(f"""
    <div class="versiculo">
    <div style="font-size:.68rem;color:#c9a84c;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">✝ Versículo del día</div>
    "{texto}"<br><strong>— {referencia}</strong>
    </div>""", unsafe_allow_html=True)

# ── DASHBOARD ─────────────────────────────────────────────────────────────────
if pagina == "📊 Dashboard":
    st.markdown('<div class="section-title">Panel Principal</div>', unsafe_allow_html=True)
    try:
        materias = get_materias()
        iconos   = ["📖","📜","👑","🕊️","🔥","⚓","🌿","🗝️","🛡️","📣"]
        if not materias:
            st.info("No hay materias aún. Ve a **Crear Materia** para empezar.")
        else:
            cols = st.columns(min(3, len(materias)))
            for i, materia in enumerate(materias):
                df       = get_estudiantes(materia)
                total    = len(df)
                promedio = round(df["Nota"].mean(), 1) if total > 0 else 0
                maxnota  = int(df["Nota"].max()) if total > 0 else 0
                with cols[i % min(3, len(materias))]:
                    st.markdown(f"""
                    <div class="materia-card">
                      <h3>{iconos[i%len(iconos)]} {materia}</h3>
                      <div class="stats">
                        <div><div style="font-size:.72rem;color:#6b7280;text-transform:uppercase;letter-spacing:1px">Estudiantes</div><div class="stat-val">{total}</div></div>
                        <div><div style="font-size:.72rem;color:#6b7280;text-transform:uppercase;letter-spacing:1px">Promedio</div><div class="stat-val">{promedio}</div></div>
                        <div><div style="font-size:.72rem;color:#6b7280;text-transform:uppercase;letter-spacing:1px">Más alta</div><div class="stat-val">{maxnota}</div></div>
                      </div>
                    </div>""", unsafe_allow_html=True)
    except Exception as e:
        st.error(f"Error conectando con Google Sheets: {e}")

# ── VER MATERIA ───────────────────────────────────────────────────────────────
elif pagina == "📖 Ver Materia":
    try:
        materias = get_materias()
        if not materias:
            st.warning("No hay materias. Crea una primero.")
        else:
            materia_sel = st.selectbox("Selecciona una materia", materias)
            df = get_estudiantes(materia_sel)
            st.markdown(f'<div class="section-title">📖 {materia_sel}</div>', unsafe_allow_html=True)

            total     = len(df)
            prom      = round(df["Nota"].mean(), 1) if total > 0 else 0
            maxn      = int(df["Nota"].max()) if total > 0 else 0
            aprobados = len(df[df["Nota"] >= 70]) if total > 0 else 0

            c1, c2, c3, c4 = st.columns(4)
            for col, label, val, sub in zip(
                [c1,c2,c3,c4],
                ["Estudiantes","Promedio","Nota más alta","Aprobados"],
                [total, prom, maxn, aprobados],
                ["inscritos","del grupo","registrada",f"de {total}"]
            ):
                col.markdown(f"""
                <div class="metric-box">
                  <div class="label">{label}</div>
                  <div class="value">{val}</div>
                  <div class="sub">{sub}</div>
                </div>""", unsafe_allow_html=True)

            st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)

            with st.expander("➕ Agregar / Actualizar Estudiante", expanded=False):
                col1, col2 = st.columns(2)
                with col1:
                    nuevo_nombre = st.text_input("Nombre completo")
                with col2:
                    nueva_nota = st.slider("Nota", 0, 100, 75)
                if st.button("💾 Guardar Estudiante"):
                    if nuevo_nombre.strip():
                        letra  = nota_a_letra(nueva_nota)
                        nueva  = pd.DataFrame([{"Nombre": nuevo_nombre.strip(), "Nota": nueva_nota,
                                                "Fecha": datetime.now().strftime("%d/%m/%Y"), "Letra": letra}])
                        df = df[df["Nombre"].str.lower() != nuevo_nombre.strip().lower()]
                        df = pd.concat([df, nueva], ignore_index=True)
                        with st.spinner("Guardando en Google Sheets…"):
                            guardar_estudiantes(materia_sel, df)
                        st.success(f"✅ {nuevo_nombre} guardado · Letra {letra}")
                        st.rerun()
                    else:
                        st.error("Escribe un nombre.")

            buscar = st.text_input("🔍 Buscar estudiante", placeholder="Escribe un nombre…")
            df_vis = df[df["Nombre"].str.contains(buscar, case=False, na=False)] if buscar else df

            if df_vis.empty:
                st.info("No hay estudiantes en esta materia aún.")
            else:
                st.dataframe(df_vis.style.format({"Nota": "{:.0f}"}),
                             use_container_width=True, hide_index=True)
                st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)
                a_eliminar = st.selectbox("Eliminar estudiante", ["— selecciona —"] + df["Nombre"].tolist())
                if a_eliminar != "— selecciona —":
                    if st.button(f"🗑️ Eliminar a {a_eliminar}"):
                        df = df[df["Nombre"] != a_eliminar]
                        with st.spinner("Guardando cambios…"):
                            guardar_estudiantes(materia_sel, df)
                        st.success(f"'{a_eliminar}' eliminado.")
                        st.rerun()

            st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)
            if not df.empty:
                fig = grafica(df, materia_sel)
                if fig: st.pyplot(fig, use_container_width=True)

            with st.expander("⚠️ Zona de peligro — Eliminar materia", expanded=False):
                st.warning("Esta acción eliminará la materia y todos sus estudiantes.")
                if st.button("🗑️ Eliminar esta materia"):
                    if eliminar_materia(materia_sel):
                        st.success("Materia eliminada."); st.rerun()
                    else:
                        st.error("No puedes eliminar la única materia.")
    except Exception as e:
        st.error(f"Error: {e}")

# ── CREAR MATERIA ─────────────────────────────────────────────────────────────
elif pagina == "➕ Crear Materia":
    st.markdown('<div class="section-title">➕ Nueva Materia</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="versiculo">
    Sugerencias: Teología Sistemática · Hermenéutica · Escatología ·
    Liderazgo Cristiano · Evangelismo · Homilética · Ética Cristiana
    </div>""", unsafe_allow_html=True)
    nombre_materia = st.text_input("Nombre de la materia", placeholder="ej: Hermenéutica")
    if st.button("✅ Crear Materia"):
        if nombre_materia.strip():
            with st.spinner("Creando materia en Google Sheets…"):
                if crear_materia(nombre_materia.strip()):
                    st.success(f"✅ Materia '{nombre_materia}' creada."); st.balloons()
                else:
                    st.warning("Ya existe una materia con ese nombre.")
        else:
            st.error("Escribe el nombre de la materia.")
    st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)
    st.markdown("**Materias actuales:**")
    try:
        for m in get_materias():
            st.markdown(f"- 📖 {m}")
    except:
        st.info("Cargando…")

# ── ESTADÍSTICAS ──────────────────────────────────────────────────────────────
elif pagina == "📈 Estadísticas":
    st.markdown('<div class="section-title">📈 Estadísticas Generales</div>', unsafe_allow_html=True)
    try:
        resumen = []
        for m in get_materias():
            df = get_estudiantes(m)
            if not df.empty:
                resumen.append({
                    "Materia": m, "Estudiantes": len(df),
                    "Promedio": round(df["Nota"].mean(), 1),
                    "Más alta": int(df["Nota"].max()),
                    "Más baja": int(df["Nota"].min()),
                    "Aprobados": len(df[df["Nota"] >= 70]),
                })
        if not resumen:
            st.info("Aún no hay datos registrados.")
        else:
            df_res = pd.DataFrame(resumen)
            st.dataframe(df_res, use_container_width=True, hide_index=True)
            st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)
            fig, ax = plt.subplots(figsize=(10, 4))
            fig.patch.set_facecolor("#0d1b3e"); ax.set_facecolor("#0a1528")
            bars = ax.bar([m[:18] for m in df_res["Materia"]], df_res["Promedio"],
                          color="#c9a84c", edgecolor="#0d1b3e", linewidth=1)
            ax.axhline(70, color="#fca5a5", linestyle="--", linewidth=1, label="Mínimo (70)")
            ax.set_title("Promedio por Materia", color="#c9a84c", fontsize=12, fontweight="bold")
            ax.tick_params(colors="#a0aec0"); ax.spines[:].set_color("#1e2d4e"); ax.set_ylim(0, 110)
            ax.legend(facecolor="#0a1528", labelcolor="#a0aec0", fontsize=8)
            for bar, val in zip(bars, df_res["Promedio"]):
                ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+1,
                        str(val), ha="center", color="white", fontsize=9, fontweight="bold")
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=20, ha="right", color="#a0aec0")
            plt.tight_layout(); st.pyplot(fig, use_container_width=True)
    except Exception as e:
        st.error(f"Error: {e}")

# ── EXPORTAR ──────────────────────────────────────────────────────────────────
elif pagina == "📥 Exportar Excel":
    st.markdown('<div class="section-title">📥 Exportar Reporte</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="versiculo">
    El archivo incluye diseño institucional, encabezados en azul marino y dorado,
    y promedio automático por materia.
    </div>""", unsafe_allow_html=True)
    try:
        materias = get_materias()
        for m in materias:
            df = get_estudiantes(m)
            c1, c2, c3 = st.columns([3,1,1])
            c1.markdown(f"📖 **{m}**")
            c2.markdown(f"{len(df)} estudiantes")
            c3.markdown(f"Prom: **{round(df['Nota'].mean(),1) if len(df)>0 else 0}**")
        st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)
        if st.button("📊 Generar Excel"):
            with st.spinner("Generando reporte…"):
                excel_data = exportar_excel()
            st.download_button(
                label="⬇️ Descargar Reporte Excel Completo",
                data=excel_data,
                file_name=f"Reporte_FuenteDeGracia_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"Error: {e}")
